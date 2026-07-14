"""Au Report Automation App.

PySide-based workflow app for one Au SEM/EDS report sample. The UI style follows
`Reference file.txt` from the `reference-file-for-codex` branch: a QMainWindow
with numbered QTabWidget workflow steps, grouped controls, status labels, tables,
and a log panel.
"""

from __future__ import annotations

import csv
import os
from difflib import SequenceMatcher
import re
import subprocess
import sys
import time
from copy import copy
from dataclasses import dataclass
from glob import glob
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from PySide6.QtCore import Qt
    from PySide6.QtGui import QPixmap
    from PySide6.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QComboBox,
        QFileDialog,
        QFrame,
        QGroupBox,
        QHBoxLayout,
        QHeaderView,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QProgressDialog,
        QPushButton,
        QTableWidget,
        QTableWidgetItem,
        QTabWidget,
        QVBoxLayout,
        QWidget,
    )
except ImportError:
    try:
        from PySide2.QtCore import Qt
        from PySide2.QtGui import QPixmap
        from PySide2.QtWidgets import (
            QApplication,
            QAbstractItemView,
            QComboBox,
            QFileDialog,
            QFrame,
            QGroupBox,
            QHBoxLayout,
            QHeaderView,
            QLabel,
            QLineEdit,
            QMainWindow,
            QMessageBox,
            QPlainTextEdit,
            QProgressDialog,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QTabWidget,
            QVBoxLayout,
            QWidget,
        )
    except ImportError as exc:
        raise SystemExit(
            "PySide6 is required to run the app. Install it with:\n"
            "    pip install PySide6\n"
            "or install PySide2 if PySide6 is not available."
        ) from exc


def _qt_attr(group_name: str, attr_name: str, fallback_name: str):
    group = getattr(Qt, group_name, None)
    if group is not None and hasattr(group, attr_name):
        return getattr(group, attr_name)
    return getattr(Qt, fallback_name)


def _class_enum(class_object, group_name: str, attr_name: str, fallback_name: str):
    group = getattr(class_object, group_name, None)
    if group is not None and hasattr(group, attr_name):
        return getattr(group, attr_name)
    return getattr(class_object, fallback_name)


ALIGN_CENTER = _qt_attr("AlignmentFlag", "AlignCenter", "AlignCenter")
SMOOTH_TRANSFORMATION = _qt_attr("TransformationMode", "SmoothTransformation", "SmoothTransformation")
HEADER_STRETCH = _class_enum(QHeaderView, "ResizeMode", "Stretch", "Stretch")
NO_EDIT_TRIGGERS = _class_enum(QAbstractItemView, "EditTrigger", "NoEditTriggers", "NoEditTriggers")

ELEMENTS_BY_SAMPLE_TYPE = {
    "Au+Ag": ("Au", "Ag"),
    "Au+Ag+Cu": ("Au", "Ag", "Cu"),
    "Au+Ag+Cu+Hg": ("Au", "Ag", "Cu", "Hg"),
    "Au+Ag+Hg": ("Au", "Ag", "Hg"),
}
SAMPLE_TYPE_TO_SCRIPT_KEY = {
    "Au+Ag": "1",
    "Au+Ag+Cu": "2",
    "Au+Ag+Cu+Hg": "3",
    "Au+Ag+Hg": "4",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg"}
RESIZED_FOLDER_NAME = "resizedtosmallest"
RAW_EXPORT_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1", "utf-16")
NUMERIC_TEXT_PATTERN = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?$")
SUMMARY_LABEL_COLUMN_OFFSET = -1
AVERAGE_LABEL = "Average"
MINIMUM_LABEL = "Minimum"
REPORT_ROW_HEIGHT = 45
REPORT_COLUMN_WIDTH = 8.43
REPORT_FONT_SIZE = 11
DISPLAY_IMAGE_SIZE_CM = 1.58
PIXELS_PER_CM_AT_96_DPI = 96 / 2.54
TARGET_IMAGE_WIDTH_PX = round(DISPLAY_IMAGE_SIZE_CM * PIXELS_PER_CM_AT_96_DPI)
TARGET_IMAGE_HEIGHT_PX = round(DISPLAY_IMAGE_SIZE_CM * PIXELS_PER_CM_AT_96_DPI)
MAX_EMBEDDED_IMAGE_PIXELS = 300
JPEG_QUALITY = 85
SAMPLE_TYPE_WARNING_THRESHOLD_WT = 2.0
MIN_NORMALIZED_TOTAL_WT = 85.0


@dataclass(frozen=True)
class AppReportLayout:
    """Image report workbook layout used directly by the app."""

    sample_type: str
    headers: tuple[str, ...]
    reflected_light_column: int
    sem_column: int


APP_REPORT_LAYOUTS = {
    "Au+Ag": AppReportLayout(
        sample_type="Au+Ag",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=4,
        sem_column=5,
    ),
    "Au+Ag+Cu": AppReportLayout(
        sample_type="Au+Ag+Cu",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Cu (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=5,
        sem_column=6,
    ),
    "Au+Ag+Cu+Hg": AppReportLayout(
        sample_type="Au+Ag+Cu+Hg",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Cu (Wt%)", "Hg (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=6,
        sem_column=7,
    ),
    "Au+Ag+Hg": AppReportLayout(
        sample_type="Au+Ag+Hg",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Hg (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=5,
        sem_column=6,
    ),
}


@dataclass
class AppState:
    sample_type: str = "Au+Ag"
    raw_export_path: Path | None = None
    sample_dir: Path | None = None
    sample_name: str = ""
    data_workbook_path: Path | None = None
    microscope_image_list_dir: Path | None = None
    microscope_resized_dir: Path | None = None
    sem_image_dir: Path | None = None
    sem_resized_dir: Path | None = None
    report_workbook_path: Path | None = None
    final_workbook_path: Path | None = None


def normalize_header(value) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def find_header_index(headers: list[str], wanted_header: str) -> int:
    wanted = normalize_header(wanted_header)
    for index, header in enumerate(headers):
        if normalize_header(header) == wanted:
            return index
    raise ValueError(f"Could not find required header {wanted_header!r}.")


def coerce_raw_value(value: str):
    """Convert numeric-looking export text to Excel numbers while keeping labels as text."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if NUMERIC_TEXT_PATTERN.match(text):
        number = float(text)
        return int(number) if number.is_integer() else number
    return value


def coerce_data_row(row: list[str]) -> list[object]:
    """Convert one non-header raw export row to Excel-friendly values."""
    return [coerce_raw_value(value) for value in row]


def read_full_export_csv(csv_path: Path) -> list[list[str]]:
    """Read a .full export CSV-like file into rows."""
    last_error: UnicodeDecodeError | None = None
    for encoding in RAW_EXPORT_ENCODINGS:
        try:
            with csv_path.open("r", encoding=encoding, newline="") as file:
                sample = file.read(4096)
                file.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return [row for row in csv.reader(file, dialect)]
        except UnicodeDecodeError as error:
            last_error = error

    raise UnicodeDecodeError(
        last_error.encoding if last_error else "unknown",
        last_error.object if last_error else b"",
        last_error.start if last_error else 0,
        last_error.end if last_error else 0,
        f"Could not decode {csv_path} using: {', '.join(RAW_EXPORT_ENCODINGS)}",
    )


def write_rows(worksheet, rows: list[list[str]]) -> None:
    for row in rows:
        worksheet.append(row)


def write_raw_rows(worksheet, rows: list[list[str]]) -> None:
    """Write raw export rows while converting numeric text to real Excel numbers."""
    if not rows:
        return
    worksheet.append(rows[0])
    for row in rows[1:]:
        worksheet.append(coerce_data_row(row))


def append_normalized_summary_formulas(
    worksheet,
    first_data_row: int,
    last_data_row: int,
    normalized_column: int,
    element_start_column: int,
    element_count: int,
) -> None:
    """Add average formulas for normalized/element columns and minimum for Normalized."""
    if last_data_row < first_data_row:
        return

    average_row = last_data_row + 1
    minimum_row = last_data_row + 2
    label_column = max(1, normalized_column + SUMMARY_LABEL_COLUMN_OFFSET)
    worksheet.cell(row=average_row, column=label_column, value=AVERAGE_LABEL)
    worksheet.cell(row=minimum_row, column=label_column, value=MINIMUM_LABEL)

    summary_columns = [normalized_column] + list(range(element_start_column, element_start_column + element_count))
    for column_number in summary_columns:
        column_letter = get_column_letter(column_number)
        worksheet.cell(
            row=average_row,
            column=column_number,
            value=f"=AVERAGE({column_letter}{first_data_row}:{column_letter}{last_data_row})",
        )
    normalized_letter = get_column_letter(normalized_column)
    worksheet.cell(
        row=minimum_row,
        column=normalized_column,
        value=f"=MIN({normalized_letter}{first_data_row}:{normalized_letter}{last_data_row})",
    )

    for row_number in (average_row, minimum_row):
        for column_number in range(label_column, element_start_column + element_count):
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if column_number in summary_columns:
                cell.number_format = "0.00"


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def create_data_workbook(raw_export_path: Path, sample_type: str) -> Path:
    """Create Output File 1 with Raw Data and Au normalized sheets."""
    rows = read_full_export_csv(raw_export_path)
    if not rows:
        raise ValueError(f"No rows were found in {raw_export_path}.")

    headers = rows[0]
    au_index = find_header_index(headers, "Au (Wt%)")
    elements = ELEMENTS_BY_SAMPLE_TYPE[sample_type]
    element_indices = {element: find_header_index(headers, f"{element} (Wt%)") for element in elements}

    workbook = Workbook()
    raw_sheet = workbook.active
    raw_sheet.title = "Raw Data"
    write_raw_rows(raw_sheet, rows)

    au_sheet = workbook.create_sheet("Au")
    au_sheet.append(headers)
    for row in rows[1:]:
        coerced_row = coerce_data_row(row)
        try:
            au_value = float(coerced_row[au_index] or 0)
        except (ValueError, IndexError):
            au_value = 0
        if au_value > 0:
            au_sheet.append(coerced_row)

    existing_last_column = len(headers)
    normalized_column = existing_last_column + 3
    element_start_column = normalized_column + 2
    au_sheet.cell(row=1, column=normalized_column, value="Normalized")
    for offset, element in enumerate(elements):
        au_sheet.cell(row=1, column=element_start_column + offset, value=element)

    for row_number in range(2, au_sheet.max_row + 1):
        source_refs = [f"{get_column_letter(element_indices[element] + 1)}{row_number}" for element in elements]
        normalized_cell = f"{get_column_letter(normalized_column)}{row_number}"
        normalized_formula_cell = au_sheet.cell(row=row_number, column=normalized_column, value=f"=SUM({','.join(source_refs)})")
        normalized_formula_cell.number_format = "0.00"
        for offset, element in enumerate(elements):
            source_cell = f"{get_column_letter(element_indices[element] + 1)}{row_number}"
            element_formula_cell = au_sheet.cell(
                row=row_number,
                column=element_start_column + offset,
                value=f'=IF({normalized_cell}=0,"",{source_cell}*100/{normalized_cell})',
            )
            element_formula_cell.number_format = "0.00"

    append_normalized_summary_formulas(
        au_sheet,
        first_data_row=2,
        last_data_row=au_sheet.max_row,
        normalized_column=normalized_column,
        element_start_column=element_start_column,
        element_count=len(elements),
    )

    for sheet in (raw_sheet, au_sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=11, bold=(cell.row == 1))
        autosize_columns(sheet)
        sheet.freeze_panes = "A2"

    output_path = raw_export_path.with_suffix(".xlsx")
    workbook.save(output_path)
    return output_path


def first_summary_row(worksheet) -> int | None:
    """Return the first row that contains an app-created summary label."""
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if str(cell.value or "").strip().lower() in {AVERAGE_LABEL.lower(), MINIMUM_LABEL.lower()}:
                return cell.row
    return None


def create_others_sheet(data_workbook_path: Path) -> None:
    """Create/update Others sheet from the edited Au sheet Area column."""
    workbook = load_workbook(data_workbook_path)
    if "Au" not in workbook.sheetnames:
        raise ValueError("The data workbook does not contain an 'Au' sheet.")
    au_sheet = workbook["Au"]
    headers = [cell.value for cell in au_sheet[1]]
    area_column_index = find_header_index(headers, "Area") + 1

    if "Others" in workbook.sheetnames:
        del workbook["Others"]
    others_sheet = workbook.create_sheet("Others")
    others_sheet.cell(row=1, column=1, value="Area")
    summary_row = first_summary_row(au_sheet)
    last_area_row = (summary_row - 1) if summary_row else au_sheet.max_row
    for row_number in range(2, last_area_row + 1):
        source_cell = au_sheet.cell(row=row_number, column=area_column_index)
        target_cell = others_sheet.cell(row=row_number, column=1, value=source_cell.value)
        if source_cell.fill:
            target_cell.fill = source_cell.fill.copy()

    others_sheet.cell(row=1, column=6, value="SEM Image No.")
    others_sheet.cell(row=1, column=7, value="uScope Image No.")
    for row in others_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=11, bold=(cell.row == 1))
    autosize_columns(others_sheet)
    workbook.save(data_workbook_path)



def recognized_element_headers(headers: list[object]) -> dict[str, int]:
    """Return raw wt% column indexes for recognized chemistry elements."""
    found = {}
    for element in ("Au", "Ag", "Cu", "Hg"):
        try:
            found[element] = find_header_index(headers, f"{element} (Wt%)")
        except ValueError:
            continue
    return found


def data_rows_before_summary(worksheet) -> range:
    """Return worksheet data-row range, excluding app summary rows."""
    summary_row = first_summary_row(worksheet)
    last_row = (summary_row - 1) if summary_row else worksheet.max_row
    return range(2, last_row + 1)


def numeric_cell_value(value) -> float:
    """Return a numeric cell value, treating blanks/formula text as zero for QC calculations."""
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def alpha_signature(value: str) -> str:
    """Return lower-case alphabetic text for loose sample/folder comparison."""
    return " ".join(re.findall(r"[a-zA-Z]+", value.lower()))


def name_similarity(left: str, right: str) -> float:
    """Return a loose 0-1 similarity score ignoring numbers and punctuation."""
    left_alpha = alpha_signature(left)
    right_alpha = alpha_signature(right)
    if not left_alpha or not right_alpha:
        return 0.0
    return SequenceMatcher(None, left_alpha, right_alpha).ratio()


def validate_data_workbook_for_app(data_workbook_path: Path, sample_type: str) -> tuple[str, str]:
    """Validate the app data workbook with short pass/warning/fail messages."""
    workbook = load_workbook(data_workbook_path, data_only=False)
    if "Au" not in workbook.sheetnames:
        return "fail", "❌ Au sheet is missing."
    if "Others" not in workbook.sheetnames:
        return "warn", "⚠️ Others sheet is missing; create it before the image report."

    au_sheet = workbook["Au"]
    headers = [cell.value for cell in au_sheet[1]]
    raw_columns = recognized_element_headers(headers)
    selected_elements = ELEMENTS_BY_SAMPLE_TYPE[sample_type]
    missing = [element for element in selected_elements if element not in raw_columns]
    if missing:
        return "fail", f"❌ Missing required raw columns: {', '.join(missing)} (Wt%)."

    extra_high = []
    for element in ("Cu", "Hg"):
        if element in selected_elements or element not in raw_columns:
            continue
        for row_number in data_rows_before_summary(au_sheet):
            if numeric_cell_value(au_sheet.cell(row=row_number, column=raw_columns[element] + 1).value) > SAMPLE_TYPE_WARNING_THRESHOLD_WT:
                extra_high.append(element)
                break
    if extra_high:
        suggested_elements = [element for element in ("Au", "Ag", "Cu", "Hg") if element in selected_elements or element in extra_high]
        suggested = "+".join(suggested_elements)
        return "fail", f"❌ Selected {sample_type}, but {', '.join(extra_high)} exceeds {SAMPLE_TYPE_WARNING_THRESHOLD_WT:g} wt%. Check sample type ({suggested})."

    low_normalized_rows = []
    normalized_sum_fail_rows = []
    normalized_percentages_by_element = {element: [] for element in selected_elements}
    for row_number in data_rows_before_summary(au_sheet):
        raw_values = [numeric_cell_value(au_sheet.cell(row=row_number, column=raw_columns[element] + 1).value) for element in selected_elements]
        normalized_total = sum(raw_values)
        if normalized_total < MIN_NORMALIZED_TOTAL_WT:
            low_normalized_rows.append(row_number)
        if normalized_total > 0:
            normalized_values = [value * 100 / normalized_total for value in raw_values]
            if sum(normalized_values) != 100.0:
                # Floating-point math can make exact equality difficult, so report only meaningful drift.
                if abs(sum(normalized_values) - 100.0) > 1e-9:
                    normalized_sum_fail_rows.append(row_number)
            for element, normalized_value in zip(selected_elements, normalized_values):
                normalized_percentages_by_element[element].append(normalized_value)

    if low_normalized_rows:
        preview = ", ".join(str(row) for row in low_normalized_rows[:5])
        return "fail", f"❌ Normalized total is below {MIN_NORMALIZED_TOTAL_WT:g} wt% in Au rows: {preview}."
    if normalized_sum_fail_rows:
        preview = ", ".join(str(row) for row in normalized_sum_fail_rows[:5])
        return "fail", f"❌ Normalized chemistry does not sum to 100 in Au rows: {preview}."

    average_sum = 0.0
    for values in normalized_percentages_by_element.values():
        if values:
            average_sum += sum(values) / len(values)
    if normalized_percentages_by_element and abs(average_sum - 100.0) > 1e-9:
        return "fail", f"❌ Average normalized chemistry sums to {average_sum:.6f}, not 100."

    area_status = validate_area_alignment(data_workbook_path)
    if area_status[0] != "pass":
        return area_status
    return "pass", "✅ Data workbook checks passed."


def validate_area_alignment(data_workbook_path: Path) -> tuple[str, str]:
    """Check that Others Area rows match edited Au rows."""
    workbook = load_workbook(data_workbook_path, data_only=False)
    if "Au" not in workbook.sheetnames or "Others" not in workbook.sheetnames:
        return "warn", "⚠️ Au or Others sheet is missing."
    au_sheet = workbook["Au"]
    others_sheet = workbook["Others"]
    au_count = len(list(data_rows_before_summary(au_sheet)))
    others_count = max(0, others_sheet.max_row - 1)
    if au_count != others_count:
        return "warn", f"⚠️ Area rows ({others_count}) do not match Au rows ({au_count})."
    return "pass", "✅ Area row count matches Au rows."


def validate_image_folder_names(raw_export_path: Path | None, microscope_folder: Path | None, sem_folder: Path | None) -> tuple[str, str]:
    """Warn if selected image folder names do not resemble the raw sample name."""
    if raw_export_path is None or microscope_folder is None or sem_folder is None:
        return "warn", "⚠️ Select raw, microscope, and SEM folders before folder-name QC."
    sample_name = raw_export_path.stem
    microscope_score = name_similarity(sample_name, microscope_folder.parent.parent.name if microscope_folder.name == "image_list" else microscope_folder.name)
    sem_score = name_similarity(sample_name, sem_folder.name)
    issues = []
    if microscope_score < 0.20:
        issues.append(f"microscope folder name match is low ({microscope_score:.0%})")
    if sem_score < 0.45:
        issues.append(f"SEM folder name match is low ({sem_score:.0%})")
    if issues:
        return "warn", "⚠️ " + "; ".join(issues) + ". Confirm folders before continuing."
    return "pass", "✅ Image folder names look consistent with the raw file."


def validate_final_outputs(final_workbook_path: Path | None) -> tuple[str, str]:
    """Check basic final workbook output existence."""
    if final_workbook_path is None or not final_workbook_path.exists():
        return "fail", "❌ Final workbook was not found."
    try:
        workbook = load_workbook(final_workbook_path, read_only=True, data_only=False)
    except Exception as exc:
        return "fail", f"❌ Could not open final workbook: {exc}"
    if "Organized Blocks" not in workbook.sheetnames:
        return "fail", "❌ Organized Blocks sheet is missing."
    return "pass", "✅ Final workbook has Organized Blocks."

def prepare_image_names(folder: Path) -> None:
    """Apply the existing image-renaming behavior before resize."""
    for image_path in folder.iterdir():
        if image_path.is_file() and image_path.suffix.lower() == ".jpeg":
            image_path.rename(image_path.with_suffix(".jpg"))

    for image_path in folder.iterdir():
        if image_path.is_file() and image_path.name.startswith("Electron Image "):
            new_name = image_path.name.replace("Electron Image ", "")
            stem, suffix = new_name.rsplit(".", 1)
            image_path.rename(image_path.with_name(f"{stem.zfill(3)}.{suffix}"))


def resize_images_to_smallest(folder: Path) -> Path:
    """Crop all JPG images in a folder to a shared smallest size."""
    try:
        import cv2
    except ImportError as exc:
        raise RuntimeError("OpenCV is required for image resizing. Install it with: pip install opencv-python") from exc

    folder = Path(folder)
    prepare_image_names(folder)
    output_folder = folder / RESIZED_FOLDER_NAME
    output_folder.mkdir(exist_ok=True)

    image_paths = sorted(Path(path) for path in glob(str(folder / "*.jpg")))
    if not image_paths:
        raise ValueError(f"No .jpg images were found in {folder}.")

    widths = []
    heights = []
    for image_path in image_paths:
        image = cv2.imread(str(image_path))
        if image is None:
            continue
        height, width = image.shape[:2]
        widths.append(width)
        heights.append(height)

    if not widths or not heights:
        raise ValueError(f"No readable .jpg images were found in {folder}.")

    crop_width = min(widths)
    crop_height = int(0.75 * crop_width) if crop_width > min(heights) else int(crop_width / 1.33)
    (output_folder / f"NEW Max Size is {crop_width} {crop_height} .txt").write_text("")

    for image_path in image_paths:
        image = cv2.imread(str(image_path))
        if image is None:
            continue
        height, width = image.shape[:2]
        mid_x, mid_y = width // 2, height // 2
        half_width, half_height = crop_width // 2, crop_height // 2
        cropped = image[mid_y - half_height:mid_y + half_height, mid_x - half_width:mid_x + half_width]
        cv2.imwrite(str(output_folder / f"{image_path.stem}-resized.jpg"), cropped)
    return output_folder


def open_path(path: Path) -> None:
    """Open a file/folder in the OS default application."""
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


def default_intermediate_report_path(sample_dir: Path, sample_name: str) -> Path:
    """Return the app's default intermediate/final report workbook path."""
    return sample_dir / f"{sample_name}_Inter.xlsx"


def default_finished_report_path(intermediate_path: Path) -> Path:
    """Return the finished workbook path for an intermediate workbook."""
    if intermediate_path.stem.endswith("_Inter"):
        return intermediate_path.with_name(f"{intermediate_path.stem[:-len('_Inter')]}_Final.xlsx")
    return intermediate_path.with_name(f"{intermediate_path.stem}_Final.xlsx")


def copy_area_column_to_report_worksheet(report_worksheet, data_workbook_path: Path, destination_column: int) -> None:
    """Copy Area from the app-created Others sheet into the report workbook.

    This local app copy avoids depending on a helper name inside
    `insert_au_report_images.py`, so users only need the app update for this fix.
    """
    data_workbook = load_workbook(data_workbook_path)
    if "Others" not in data_workbook.sheetnames:
        raise ValueError(f"The data workbook does not contain an 'Others' sheet: {data_workbook_path}")

    source_worksheet = data_workbook["Others"]
    headers = [cell.value for cell in source_worksheet[1]]
    source_column = find_header_index(headers, "Area") + 1

    header_cell = report_worksheet.cell(row=1, column=destination_column, value="Area")
    header_cell.font = Font(size=REPORT_FONT_SIZE, bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    report_worksheet.column_dimensions[get_column_letter(destination_column)].width = REPORT_COLUMN_WIDTH

    for row_number in range(2, source_worksheet.max_row + 1):
        source_cell = source_worksheet.cell(row=row_number, column=source_column)
        destination_cell = report_worksheet.cell(row=row_number, column=destination_column, value=source_cell.value)
        destination_cell.font = Font(size=REPORT_FONT_SIZE, bold=False)
        destination_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        destination_cell.fill = copy(source_cell.fill)
        report_worksheet.row_dimensions[row_number].height = REPORT_ROW_HEIGHT


def first_filename_number(path: Path) -> int:
    """Return the first number in a filename for numeric image sorting."""
    match = re.search(r"\d+", path.stem)
    if not match:
        return 10**12
    return int(match.group())


def sorted_report_image_files(folder: Path) -> list[Path]:
    """Return report images sorted from lowest filename number to highest."""
    images = [path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS]
    return sorted(images, key=lambda path: (first_filename_number(path), path.name.lower()))


def set_report_dimensions(worksheet, max_row: int, max_column: int) -> None:
    """Apply fixed report row height and column width."""
    for row_number in range(1, max_row + 1):
        worksheet.row_dimensions[row_number].height = REPORT_ROW_HEIGHT
    for column_number in range(1, max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = REPORT_COLUMN_WIDTH


def apply_report_text_format(worksheet, max_row: int, max_column: int) -> None:
    """Apply report text formatting with bold headers only."""
    for row_number in range(1, max_row + 1):
        for column_number in range(1, max_column + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            cell.font = Font(size=REPORT_FONT_SIZE, bold=(row_number == 1))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def prepare_report_image(image_path: Path, temporary_folder: Path, image_role: str, row_number: int) -> Path:
    """Create a compressed temporary image copy for Excel embedding.

    The temporary filename includes the image role and row number so microscope
    and SEM images with the same source filename can never overwrite each other.
    """
    try:
        from PIL import Image as PillowImage
    except ImportError as exc:
        raise RuntimeError("Pillow is required to create the image report workbook. Install it with: pip install Pillow") from exc

    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", image_path.stem)[:60]
    output_path = temporary_folder / f"{image_role}_{row_number:04d}_{safe_stem}_au_report.jpg"
    with PillowImage.open(image_path) as source_image:
        embedded_image = source_image.convert("RGB")
        embedded_image.thumbnail((MAX_EMBEDDED_IMAGE_PIXELS, MAX_EMBEDDED_IMAGE_PIXELS), PillowImage.Resampling.LANCZOS)
        embedded_image.save(output_path, format="JPEG", quality=JPEG_QUALITY, optimize=True)
    return output_path


def add_report_image_to_cell(
    worksheet,
    image_path: Path,
    row_number: int,
    column_number: int,
    temporary_folder: Path,
    image_role: str,
) -> None:
    """Insert one image into the report workbook at the requested display size."""
    from openpyxl.drawing.image import Image as ExcelImage

    embedded_image_path = prepare_report_image(image_path, temporary_folder, image_role, row_number)
    image = ExcelImage(str(embedded_image_path))
    image.width = TARGET_IMAGE_WIDTH_PX
    image.height = TARGET_IMAGE_HEIGHT_PX
    worksheet.add_image(image, f"{get_column_letter(column_number)}{row_number}")


def create_app_image_workbook(
    layout: AppReportLayout,
    reflected_images: list[Path],
    sem_images: list[Path],
    temporary_folder: Path,
) -> Workbook:
    """Create the image report workbook directly in the app.

    This mirrors the standalone image insertion behavior while avoiding fragile
    function-signature coupling to `insert_au_report_images.py`.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = layout.sample_type.replace("+", "_")

    image_count = max(len(reflected_images), len(sem_images))
    max_column = len(layout.headers)
    final_row = image_count + 1
    set_report_dimensions(worksheet, final_row, max_column)
    apply_report_text_format(worksheet, final_row, max_column)

    for column_number, header in enumerate(layout.headers, start=1):
        header_cell = worksheet.cell(row=1, column=column_number, value=header)
        header_cell.font = Font(size=REPORT_FONT_SIZE, bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index in range(image_count):
        row_number = index + 2
        number_cell = worksheet.cell(row=row_number, column=1, value=index + 1)
        number_cell.font = Font(size=REPORT_FONT_SIZE, bold=False)
        number_cell.alignment = Alignment(horizontal="center", vertical="center")

        if index < len(reflected_images):
            add_report_image_to_cell(
                worksheet,
                reflected_images[index],
                row_number,
                layout.reflected_light_column,
                temporary_folder,
                "microscope",
            )
        if index < len(sem_images):
            add_report_image_to_cell(worksheet, sem_images[index], row_number, layout.sem_column, temporary_folder, "sem")

    worksheet.freeze_panes = "A2"
    return workbook


def create_intermediate_report_workbook(
    sample_type: str,
    reflected_folder: Path,
    sem_folder: Path,
    data_workbook_path: Path,
    output_path: Path,
) -> Path:
    """Create the image report workbook using the already-known app paths."""
    layout = APP_REPORT_LAYOUTS[sample_type]
    reflected_images = sorted_report_image_files(reflected_folder)
    sem_images = sorted_report_image_files(sem_folder)
    if not reflected_images and not sem_images:
        raise ValueError("No .jpg or .jpeg images were found in the resized microscope or SEM folders.")

    with TemporaryDirectory() as temporary_directory:
        workbook = create_app_image_workbook(layout, reflected_images, sem_images, Path(temporary_directory))
        worksheet = workbook.active
        copy_area_column_to_report_worksheet(worksheet, data_workbook_path, layout.sem_column + 1)
        workbook.save(output_path)
    return output_path


def finish_report_from_app(sample_type: str, report_path: Path, data_workbook_path: Path, output_path: Path) -> Path:
    """Run the Excel finishing script from app state without repeated prompts."""
    try:
        from AuReportAutomation.finish_au_report import SAMPLE_FINISH_LAYOUTS, finish_report_workbook
    except ModuleNotFoundError:
        from finish_au_report import SAMPLE_FINISH_LAYOUTS, finish_report_workbook  # type: ignore[no-redef]

    layout = SAMPLE_FINISH_LAYOUTS[SAMPLE_TYPE_TO_SCRIPT_KEY[sample_type]]
    return finish_report_workbook(layout, report_path, data_workbook_path, output_path)


def find_existing_final_workbook(state: AppState) -> Path | None:
    """Find the final workbook known or inferable from app state."""
    candidates: list[Path] = []
    if state.final_workbook_path is not None:
        candidates.append(state.final_workbook_path)
    if state.report_workbook_path is not None:
        if state.report_workbook_path.name.endswith("_Final.xlsx"):
            candidates.append(state.report_workbook_path)
        candidates.append(default_finished_report_path(state.report_workbook_path))
    if state.sample_dir is not None and state.sample_name:
        candidates.append(state.sample_dir / f"{state.sample_name}_Final.xlsx")

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_existing_data_workbook(state: AppState) -> Path | None:
    """Find an existing Output File 1 workbook from selected raw/sample state."""
    candidates: list[Path] = []
    if state.data_workbook_path is not None:
        candidates.append(state.data_workbook_path)
    if state.raw_export_path is not None:
        candidates.append(state.raw_export_path.with_suffix(".xlsx"))
    if state.sample_dir is not None and state.sample_name:
        candidates.append(state.sample_dir / f"{state.sample_name}.xlsx")

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


class AuReportAutomationApp(QMainWindow):
    """PySide workflow app for the Au report automation."""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Au Report Automation App")
        self.resize(1280, 820)
        self.state = AppState()

        self.tabs = QTabWidget()
        self.setup_tab = self._build_setup_tab()
        self.images_tab = self._build_images_tab()
        self.report_tab = self._build_report_tab()
        self.word_tab = self._build_word_tab()
        self.tabs.addTab(self.setup_tab, "1. Data Workbook")
        self.tabs.addTab(self.images_tab, "2. Resize Images")
        self.tabs.addTab(self.report_tab, "3. Excel Report")
        self.tabs.addTab(self.word_tab, "4. Word Report")
        central_widget = QWidget()
        central_layout = QVBoxLayout(central_widget)
        central_layout.addWidget(self._build_app_header())
        central_layout.addWidget(self.tabs, stretch=1)
        self.setCentralWidget(central_widget)
        self._apply_professional_style()

    def _build_app_header(self):
        """Build a company-style header with the exact local AMTEL logo image.

        Save the provided logo image locally as `AuReportAutomation/assets/amtel_logo.png`
        (or `.jpg` / `.gif`). The app loads that real image directly instead of
        redrawing or altering the logo in code.
        """
        header = QFrame()
        header.setObjectName("appHeader")
        header_layout = QHBoxLayout(header)

        logo_label = QLabel("AMTEL")
        logo_label.setObjectName("logoText")
        logo_label.setToolTip(
            "Place the exact logo image at AuReportAutomation/assets/amtel_logo.png "
            "or AuReportAutomation/assests/amtel_logo.png"
        )
        logo_folders = ("assets", "assests")  # support the common misspelling too
        logo_names = ("amtel_logo.png", "amtel_logo.jpg", "amtel_logo.gif")
        for folder_name in logo_folders:
            for logo_name in logo_names:
                logo_path = Path(__file__).with_name(folder_name) / logo_name
                if logo_path.exists():
                    pixmap = QPixmap(str(logo_path))
                    if not pixmap.isNull():
                        logo_label.setPixmap(pixmap.scaledToHeight(54, SMOOTH_TRANSFORMATION))
                        logo_label.setText("")
                        break
            if not logo_label.text():
                break

        title_layout = QVBoxLayout()
        title = QLabel("Au Report Automation App")
        title.setObjectName("appTitle")
        subtitle = QLabel("AMTEL SEM/EDS reporting workflow")
        subtitle.setObjectName("appSubtitle")
        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)

        header_layout.addWidget(logo_label)
        header_layout.addLayout(title_layout)
        header_layout.addStretch(1)
        return header

    def _apply_professional_style(self) -> None:
        """Apply light, professional colors to the workflow app."""
        self.setStyleSheet(
            """
            QMainWindow { background-color: #f5f7fb; }
            QWidget { font-family: Segoe UI, Arial, sans-serif; font-size: 10pt; }
            QFrame#appHeader {
                background-color: #ffffff;
                border: 1px solid #d8e2ef;
                border-radius: 10px;
                padding: 10px;
            }
            QLabel#logoText {
                color: #0f766e;
                font-size: 28px;
                font-weight: 800;
                letter-spacing: 2px;
                padding: 4px 14px;
                border: 2px solid #99f6e4;
                border-radius: 8px;
                background-color: #ecfeff;
            }
            QLabel#appTitle { color: #0f172a; font-size: 20px; font-weight: 800; }
            QLabel#appSubtitle { color: #0f766e; font-size: 11px; font-weight: 600; }
            QTabWidget::pane { border: 1px solid #cbd5e1; background: #ffffff; }
            QTabBar::tab {
                background: #e8eef8;
                border: 1px solid #cbd5e1;
                border-bottom: none;
                padding: 8px 14px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
            }
            QTabBar::tab:selected { background: #ffffff; color: #1d4ed8; font-weight: 700; }
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #d8e2ef;
                border-radius: 8px;
                margin-top: 10px;
                padding: 12px;
                font-weight: 600;
                color: #1f2937;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
                background-color: #ffffff;
            }
            QPushButton {
                background-color: #e0f2fe;
                border: 1px solid #93c5fd;
                border-radius: 6px;
                padding: 7px 12px;
                color: #0f172a;
                font-weight: 600;
            }
            QPushButton:hover { background-color: #dbeafe; }
            QPushButton:disabled { background-color: #e5e7eb; color: #6b7280; border-color: #d1d5db; }
            QPushButton#primaryButton { background-color: #dcfce7; border-color: #86efac; }
            QPushButton#primaryButton:hover { background-color: #bbf7d0; }
            QPushButton#manualButton { background-color: #fef3c7; border-color: #facc15; }
            QPushButton#manualButton:hover { background-color: #fde68a; }
            QPushButton#wordButton { background-color: #ede9fe; border-color: #c4b5fd; }
            QPushButton#wordButton:hover { background-color: #ddd6fe; }
            QLabel { color: #334155; }
            QLineEdit, QPlainTextEdit, QTableWidget {
                background-color: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 5px;
            }
            QLineEdit { padding: 6px; }
            QPlainTextEdit { padding: 6px; color: #1f2937; }
            QHeaderView::section {
                background-color: #eff6ff;
                color: #1e3a8a;
                border: 1px solid #bfdbfe;
                padding: 6px;
                font-weight: 700;
            }
            QTableWidget::item { padding: 5px; }
            QProgressDialog { background-color: #ffffff; }
            QProgressBar {
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                text-align: center;
                background-color: #f1f5f9;
                min-height: 18px;
            }
            QProgressBar::chunk {
                background-color: #86efac;
                border-radius: 6px;
            }
            """
        )

    def _build_setup_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)

        group = QGroupBox("Sample Setup and Output File 1")
        group_layout = QVBoxLayout(group)

        sample_layout = QHBoxLayout()
        self.sample_type_combo = QComboBox()
        self.sample_type_combo.addItems(ELEMENTS_BY_SAMPLE_TYPE.keys())
        self.raw_export_edit = QLineEdit()
        self.raw_export_edit.setPlaceholderText("Select .full export / CSV file")
        self.select_raw_button = QPushButton("Select Raw Export")
        sample_layout.addWidget(QLabel("Sample type"))
        sample_layout.addWidget(self.sample_type_combo)
        sample_layout.addWidget(self.raw_export_edit, stretch=1)
        sample_layout.addWidget(self.select_raw_button)

        buttons_layout = QHBoxLayout()
        self.create_data_button = QPushButton("Create Data Workbook")
        self.create_others_button = QPushButton("Create Others Sheet After Manual Au Edits")
        self.select_data_button = QPushButton("Select Existing Data Workbook")
        self.open_data_button = QPushButton("Open Data Workbook")
        self.create_data_button.setObjectName("primaryButton")
        self.create_others_button.setObjectName("manualButton")
        self.select_data_button.setObjectName("manualButton")
        self.open_data_button.setObjectName("manualButton")
        buttons_layout.addWidget(self.create_data_button)
        buttons_layout.addWidget(self.create_others_button)
        buttons_layout.addWidget(self.select_data_button)
        buttons_layout.addWidget(self.open_data_button)
        buttons_layout.addStretch(1)

        self.data_status = QLabel("Select the .full export file to start. The containing folder becomes the sample folder.")
        group_layout.addLayout(sample_layout)
        group_layout.addLayout(buttons_layout)
        group_layout.addWidget(self.data_status)

        self.summary_table = QTableWidget(0, 2)
        self.summary_table.setHorizontalHeaderLabels(["Item", "Path / Value"])
        self.summary_table.horizontalHeader().setSectionResizeMode(HEADER_STRETCH)
        self.summary_table.verticalHeader().setVisible(False)
        self.summary_table.setEditTriggers(NO_EDIT_TRIGGERS)

        self.workflow_table = QTableWidget(0, 3)
        self.workflow_table.setHorizontalHeaderLabels(["Step", "Status", "What it needs"])
        self.workflow_table.horizontalHeader().setSectionResizeMode(HEADER_STRETCH)
        self.workflow_table.verticalHeader().setVisible(False)
        self.workflow_table.setEditTriggers(NO_EDIT_TRIGGERS)

        layout.addWidget(group)
        layout.addWidget(self.summary_table, stretch=1)
        layout.addWidget(QLabel("Workflow checklist"))
        layout.addWidget(self.workflow_table, stretch=1)
        self.app_log = QPlainTextEdit()
        self.app_log.setReadOnly(True)
        layout.addWidget(self.app_log, stretch=1)
        self._connect_setup_signals()
        return tab

    def _build_images_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        group = QGroupBox("Resize Microscope and SEM Images")
        group_layout = QVBoxLayout(group)

        self.resize_microscope_button = QPushButton("Select Microscope Parent Folder and Resize")
        self.resize_sem_button = QPushButton("Select SEM Folder and Resize")
        self.resize_microscope_button.setObjectName("primaryButton")
        self.resize_sem_button.setObjectName("primaryButton")
        self.image_status = QLabel("Resize microscope images first, then SEM images.")
        group_layout.addWidget(self.resize_microscope_button)
        group_layout.addWidget(self.resize_sem_button)
        group_layout.addWidget(self.image_status)
        layout.addWidget(group)
        layout.addStretch(1)

        self.resize_microscope_button.clicked.connect(self._resize_microscope_images)
        self.resize_sem_button.clicked.connect(self._resize_sem_images)
        return tab

    def _build_report_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        group = QGroupBox("Excel Report Workflow")
        group_layout = QVBoxLayout(group)
        self.create_report_button = QPushButton("Create Image Report Workbook")
        self.open_report_button = QPushButton("Open Image Report for Manual Edits")
        self.finish_report_button = QPushButton("Finish Excel Report and Create Organized Blocks")
        self.create_report_button.setObjectName("primaryButton")
        self.open_report_button.setObjectName("manualButton")
        self.finish_report_button.setObjectName("primaryButton")
        self.report_status = QLabel("Create the image report, make manual edits, then finish the Excel report.")
        group_layout.addWidget(self.create_report_button)
        group_layout.addWidget(self.open_report_button)
        group_layout.addWidget(self.finish_report_button)
        group_layout.addWidget(self.report_status)
        layout.addWidget(group)
        layout.addStretch(1)
        self.create_report_button.clicked.connect(self._create_image_report_workbook)
        self.open_report_button.clicked.connect(self._open_report_workbook)
        self.finish_report_button.clicked.connect(self._finish_excel_report)
        return tab

    def _build_word_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        group = QGroupBox("Word Report Workflow")
        group_layout = QVBoxLayout(group)
        self.run_word_button = QPushButton("Run Word Report Macro Launcher")
        self.open_word_button = QPushButton("Open Word Report")
        self.run_word_button.setObjectName("wordButton")
        self.open_word_button.setObjectName("manualButton")
        self.word_status = QLabel(
            "Runs the Word macro launcher using the final workbook and sample type already known by the app."
        )
        group_layout.addWidget(self.run_word_button)
        group_layout.addWidget(self.open_word_button)
        group_layout.addWidget(self.word_status)
        layout.addWidget(group)
        layout.addStretch(1)
        self.run_word_button.clicked.connect(self._run_word_report_launcher)
        self.open_word_button.clicked.connect(self._open_word_report)
        return tab

    def _connect_setup_signals(self) -> None:
        self.sample_type_combo.currentTextChanged.connect(self._set_sample_type)
        self.select_raw_button.clicked.connect(self._select_raw_export)
        self.create_data_button.clicked.connect(self._create_data_workbook)
        self.create_others_button.clicked.connect(self._create_others_sheet)
        self.select_data_button.clicked.connect(self._select_existing_data_workbook)
        self.open_data_button.clicked.connect(self._open_data_workbook)

    def _set_sample_type(self, sample_type: str) -> None:
        self.state.sample_type = sample_type
        self._refresh_summary()

    def _select_raw_export(self) -> None:
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Select .full export / CSV file",
            str(Path.home()),
            "All files (*.*);;EDS exports (*.full* *.csv *.txt)",
        )
        if not selected:
            return
        raw_path = Path(selected)
        self.state.raw_export_path = raw_path
        self.state.sample_dir = raw_path.parent
        self.state.sample_name = raw_path.stem
        self.raw_export_edit.setText(str(raw_path))
        self.data_status.setText(f"Selected raw export. Sample folder: {raw_path.parent}")
        self._append_log(f"Selected raw export: {raw_path}")
        existing_data_workbook = find_existing_data_workbook(self.state)
        if existing_data_workbook is not None:
            self.state.data_workbook_path = existing_data_workbook
            self.data_status.setText(f"Selected raw export and detected existing data workbook: {existing_data_workbook}")
            self._append_log(f"Detected existing data workbook: {existing_data_workbook}")
        self._refresh_summary()

    def _set_step_status(self, label: QLabel, level: str, message: str) -> None:
        """Show compact pass/warning/fail status without flooding the UI."""
        colors = {
            "pass": ("#dcfce7", "#166534"),
            "warn": ("#fef9c3", "#92400e"),
            "fail": ("#fee2e2", "#991b1b"),
            "info": ("#e0f2fe", "#075985"),
        }
        background, foreground = colors.get(level, colors["info"])
        label.setText(message)
        label.setStyleSheet(
            f"background-color: {background}; color: {foreground}; border-radius: 6px; padding: 6px; font-weight: 600;"
        )

    def _validate_current_data_workbook(self) -> None:
        if self.state.data_workbook_path is None:
            return
        level, message = validate_data_workbook_for_app(self.state.data_workbook_path, self.state.sample_type)
        self._set_step_status(self.data_status, level, message)
        self._append_log(message)

    def _create_data_workbook(self) -> None:
        if self.state.raw_export_path is None:
            QMessageBox.warning(self, "Missing raw export", "Select the .full export file first.")
            return
        try:
            self.state.data_workbook_path = create_data_workbook(self.state.raw_export_path, self.state.sample_type)
            self._append_log(f"Created data workbook: {self.state.data_workbook_path}")
            self._validate_current_data_workbook()
            self._refresh_summary()
            open_path(self.state.data_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Data workbook failed", str(exc))

    def _create_others_sheet(self) -> None:
        if self.state.data_workbook_path is None:
            QMessageBox.warning(self, "Missing data workbook", "Create the data workbook first.")
            return
        try:
            create_others_sheet(self.state.data_workbook_path)
            self.data_status.setText("Created/updated Others sheet after manual Au edits.")
            self._append_log(f"Created/updated Others sheet in: {self.state.data_workbook_path}")
            open_path(self.state.data_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Others sheet failed", str(exc))

    def _open_data_workbook(self) -> None:
        if self.state.data_workbook_path and self.state.data_workbook_path.exists():
            open_path(self.state.data_workbook_path)

    def _select_existing_data_workbook(self) -> None:
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Select existing data workbook with Au and Others sheets",
            str(self.state.sample_dir or Path.home()),
            "Excel workbooks (*.xlsx *.xlsm);;All files (*.*)",
        )
        if not selected:
            return
        self.state.data_workbook_path = Path(selected)
        self._append_log(f"Selected existing data workbook: {self.state.data_workbook_path}")
        self._validate_current_data_workbook()
        self._refresh_summary()

    def _resize_microscope_images(self) -> None:
        selected = QFileDialog.getExistingDirectory(self, "Select microscope parent folder")
        if not selected:
            return
        image_list_dir = Path(selected) / "grains" / "image_list"
        if not image_list_dir.exists():
            QMessageBox.warning(self, "Missing image_list", f"Could not find:\n{image_list_dir}")
            return
        try:
            self.state.microscope_image_list_dir = image_list_dir
            self.state.microscope_resized_dir = resize_images_to_smallest(image_list_dir)
            self._set_step_status(self.image_status, "pass", "✅ Microscope resize complete. Now select SEM images.")
            self._append_log(f"Created microscope resized folder: {self.state.microscope_resized_dir}")
            self._refresh_summary()
        except Exception as exc:
            QMessageBox.critical(self, "Microscope resize failed", str(exc))

    def _resize_sem_images(self) -> None:
        selected = QFileDialog.getExistingDirectory(self, "Select SEM image folder")
        if not selected:
            return
        try:
            self.state.sem_image_dir = Path(selected)
            self.state.sem_resized_dir = resize_images_to_smallest(self.state.sem_image_dir)
            level, message = validate_image_folder_names(self.state.raw_export_path, self.state.microscope_image_list_dir, self.state.sem_image_dir)
            if level == "pass":
                message = "✅ SEM resize complete. Image folder checks passed."
            self._set_step_status(self.image_status, level, message)
            self._append_log(f"Created SEM resized folder: {self.state.sem_resized_dir}")
            self._append_log(message)
            self._refresh_summary()
        except Exception as exc:
            QMessageBox.critical(self, "SEM resize failed", str(exc))

    def _create_image_report_workbook(self) -> None:
        if self.state.data_workbook_path is None:
            self.state.data_workbook_path = find_existing_data_workbook(self.state)

        missing = []
        if self.state.sample_dir is None:
            missing.append("sample folder from the selected .full export file")
        if self.state.data_workbook_path is None:
            missing.append("data workbook with Others sheet")
        if self.state.microscope_resized_dir is None:
            missing.append("resized microscope image folder")
        if self.state.sem_resized_dir is None:
            missing.append("resized SEM image folder")
        if missing:
            QMessageBox.warning(self, "Missing inputs", "Complete these steps first:\n- " + "\n- ".join(missing))
            return

        folder_level, folder_message = validate_image_folder_names(
            self.state.raw_export_path, self.state.microscope_image_list_dir, self.state.sem_image_dir
        )
        if folder_level != "pass":
            self._set_step_status(self.report_status, folder_level, folder_message)
            self._append_log(folder_message)

        output_path = default_intermediate_report_path(self.state.sample_dir, self.state.sample_name)
        try:
            self.state.report_workbook_path = create_intermediate_report_workbook(
                self.state.sample_type,
                self.state.microscope_resized_dir,
                self.state.sem_resized_dir,
                self.state.data_workbook_path,
                output_path,
            )
            self._set_step_status(self.report_status, "pass", "✅ Image report workbook created. Review/manual edit before finishing.")
            self._append_log(f"Created image report workbook: {self.state.report_workbook_path}")
            self._refresh_summary()
            open_path(self.state.report_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Image report failed", str(exc))

    def _open_report_workbook(self) -> None:
        if self.state.report_workbook_path and self.state.report_workbook_path.exists():
            open_path(self.state.report_workbook_path)
        else:
            QMessageBox.warning(self, "Missing report workbook", "Create the image report workbook first.")

    def _finish_excel_report(self) -> None:
        if self.state.report_workbook_path is None or self.state.data_workbook_path is None:
            QMessageBox.warning(
                self,
                "Missing inputs",
                "Create the image report workbook and data workbook first. Save/close the report after manual edits.",
            )
            return
        output_path = default_finished_report_path(self.state.report_workbook_path)
        progress = QProgressDialog("Preparing Excel finishing step...", None, 0, 100, self)
        progress.setWindowTitle("Au Report Automation")
        progress.setCancelButton(None)
        progress.setMinimumDuration(0)
        progress.setValue(5)
        progress.show()
        self.finish_report_button.setEnabled(False)
        self.report_status.setText("5% - Preparing Excel finishing step.")
        QApplication.processEvents()
        try:
            progress.setLabelText("25% - Excel is pasting chemistry and creating Organized Blocks...")
            progress.setValue(25)
            self.report_status.setText("25% - Excel is pasting chemistry and creating Organized Blocks...")
            QApplication.processEvents()
            self.state.final_workbook_path = finish_report_from_app(
                self.state.sample_type,
                self.state.report_workbook_path,
                self.state.data_workbook_path,
                output_path,
            )
            progress.setLabelText("90% - Final workbook created. Cleaning up intermediate workbook...")
            progress.setValue(90)
            self.report_status.setText("90% - Final workbook created. Cleaning up intermediate workbook...")
            QApplication.processEvents()
            intermediate_path = self.state.report_workbook_path
            if intermediate_path != self.state.final_workbook_path and intermediate_path.exists():
                try:
                    intermediate_path.unlink()
                    self._append_log(f"Removed intermediate workbook after final was created: {intermediate_path}")
                except OSError as cleanup_error:
                    self._append_log(f"Could not remove intermediate workbook {intermediate_path}: {cleanup_error}")
            self.state.report_workbook_path = self.state.final_workbook_path
            level, message = validate_final_outputs(self.state.final_workbook_path)
            self._set_step_status(self.report_status, level, message)
            self._append_log(f"Finished Excel report: {self.state.final_workbook_path}")
            self._append_log(message)
            self._refresh_summary()
            progress.setLabelText("100% - Finished Excel report.")
            progress.setValue(100)
            QApplication.processEvents()
            open_path(self.state.final_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Finish report failed", str(exc))
        finally:
            self.finish_report_button.setEnabled(True)
            progress.close()

    def _run_word_report_launcher(self) -> None:
        launcher_path = Path(__file__).with_name("run_au_word_report_macro.py")
        if not launcher_path.exists():
            QMessageBox.critical(self, "Missing launcher", f"Could not find:\n{launcher_path}")
            return
        final_workbook_path = find_existing_final_workbook(self.state)
        if final_workbook_path is None:
            QMessageBox.warning(self, "Missing final workbook", "Finish the Excel report first so the app knows the final workbook path.")
            return
        self.state.final_workbook_path = final_workbook_path
        self.state.report_workbook_path = final_workbook_path
        self._refresh_summary()
        try:
            subprocess.Popen(
                [
                    sys.executable,
                    str(launcher_path),
                    "--workbook",
                    str(final_workbook_path),
                    "--sample-type",
                    self.state.sample_type,
                ]
            )
            self._set_step_status(self.word_status, "pass", "✅ Word export started. Use Open Word Report after it finishes.")
            self._append_log(
                f"Started Word report macro launcher for {final_workbook_path} as {self.state.sample_type}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Word launcher failed", str(exc))

    def _open_word_report(self) -> None:
        final_workbook_path = find_existing_final_workbook(self.state)
        if final_workbook_path is None:
            QMessageBox.warning(self, "Missing final workbook", "Finish the Excel report first.")
            return
        word_path = final_workbook_path.with_suffix(".docx")
        if not word_path.exists():
            QMessageBox.warning(self, "Missing Word report", f"Could not find:\n{word_path}\n\nRun the Word export first.")
            return
        open_path(word_path)

    def _refresh_summary(self) -> None:
        rows = [
            ("Sample type", self.state.sample_type),
            ("Sample name", self.state.sample_name),
            ("Sample folder", self.state.sample_dir),
            ("Raw export", self.state.raw_export_path),
            ("Data workbook", self.state.data_workbook_path),
            ("Microscope image_list", self.state.microscope_image_list_dir),
            ("Microscope resized", self.state.microscope_resized_dir),
            ("SEM folder", self.state.sem_image_dir),
            ("SEM resized", self.state.sem_resized_dir),
            ("Report workbook", self.state.report_workbook_path),
            ("Final workbook", self.state.final_workbook_path),
        ]
        self.summary_table.setRowCount(len(rows))
        for row_number, (label, value) in enumerate(rows):
            label_item = QTableWidgetItem(str(label))
            value_item = QTableWidgetItem("" if value is None else str(value))
            label_item.setTextAlignment(ALIGN_CENTER)
            value_item.setTextAlignment(ALIGN_CENTER)
            self.summary_table.setItem(row_number, 0, label_item)
            self.summary_table.setItem(row_number, 1, value_item)
        self._refresh_workflow_checklist()

    def _refresh_workflow_checklist(self) -> None:
        if not hasattr(self, "workflow_table"):
            return

        data_ready = self.state.data_workbook_path is not None and self.state.data_workbook_path.exists()
        images_ready = (
            self.state.microscope_resized_dir is not None
            and self.state.microscope_resized_dir.exists()
            and self.state.sem_resized_dir is not None
            and self.state.sem_resized_dir.exists()
        )
        report_ready = self.state.report_workbook_path is not None and self.state.report_workbook_path.exists()
        final_ready = self.state.final_workbook_path is not None and self.state.final_workbook_path.exists()

        checklist_rows = [
            (
                "1. Data Workbook",
                "Done" if data_ready else ("Ready" if self.state.raw_export_path else "Not started"),
                "Select raw export, then create or select existing data workbook.",
            ),
            (
                "2. Resize Images",
                "Done" if images_ready else ("Ready" if self.state.raw_export_path else "Not started"),
                "Run microscope resize and SEM resize.",
            ),
            (
                "3. Excel Report",
                "Done" if final_ready else ("Ready" if data_ready and images_ready else "Needs attention"),
                "Create image report, manually edit it, then finish Excel report.",
            ),
            (
                "4. Word Report",
                "Ready" if final_ready else "Needs attention",
                "Requires the final workbook from Step 3.",
            ),
        ]

        self.workflow_table.setRowCount(len(checklist_rows))
        for row_number, row_values in enumerate(checklist_rows):
            for column_number, value in enumerate(row_values):
                item = QTableWidgetItem(value)
                item.setTextAlignment(ALIGN_CENTER)
                self.workflow_table.setItem(row_number, column_number, item)

    def _append_log(self, message: str) -> None:
        timestamp = time.strftime("%H:%M:%S")
        if hasattr(self, "app_log"):
            self.app_log.appendPlainText(f"[{timestamp}] {message}")


def main() -> None:
    app = QApplication(sys.argv)
    window = AuReportAutomationApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
