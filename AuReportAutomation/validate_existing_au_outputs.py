"""Temporary QC checker for already-created Au report output files.

This helper is intentionally separate from the main app. It is for checking older
sample outputs that were already created before the app-level QC checks are added.
It asks for:
1. the final Word report (.docx)
2. the data workbook with Raw/Au/Others sheets (.xlsx/.xlsm)

It writes a text QC report next to the selected Word file and shows a compact
Pass/Warning/Fail message.
"""

from __future__ import annotations

from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

ELEMENTS = ("Au", "Ag", "Cu", "Hg")
AVERAGE_LABEL = "Average"
MINIMUM_LABEL = "Minimum"
MIN_NORMALIZED_TOTAL_WT = 85.0
SAMPLE_TYPE_WARNING_THRESHOLD_WT = 2.0


def normalize_header(value) -> str:
    """Normalize headers so wrapped/spaced Excel headers still match."""
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def find_header(headers: list[object], wanted: str) -> int | None:
    """Return a 1-based column number for a header, or None if missing."""
    wanted_normalized = normalize_header(wanted)
    for index, header in enumerate(headers, start=1):
        if normalize_header(header) == wanted_normalized:
            return index
    return None


def first_summary_row(worksheet) -> int | None:
    """Return the first row containing an Average/Minimum summary label."""
    labels = {AVERAGE_LABEL.lower(), MINIMUM_LABEL.lower()}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if str(cell.value or "").strip().lower() in labels:
                return cell.row
    return None


def data_row_numbers(worksheet) -> list[int]:
    """Return data row numbers excluding the app-created summary rows."""
    summary_row = first_summary_row(worksheet)
    last_row = (summary_row - 1) if summary_row else worksheet.max_row
    return list(range(2, last_row + 1))


def numeric_value(value) -> float:
    """Convert an Excel cell value to float for QC calculations."""
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def detected_normalized_elements(headers: list[object]) -> list[str]:
    """Infer selected sample elements from headers after the Normalized header."""
    normalized_column = find_header(headers, "Normalized")
    if normalized_column is None:
        return []
    found = []
    for header in headers[normalized_column:]:
        text = str(header or "").strip()
        if text in ELEMENTS and text not in found:
            found.append(text)
    return found


def validate_data_workbook(workbook_path: Path) -> tuple[str, list[str]]:
    """Validate Raw/Au/Others workbook content that can be checked without final Excel."""
    messages = []
    status = "pass"
    workbook = load_workbook(workbook_path, data_only=False)

    if len(workbook.sheetnames) < 3:
        return "fail", [f"❌ Workbook has only {len(workbook.sheetnames)} sheet(s); expected Raw/Au/Others."]
    messages.append(f"✅ Raw sheet considered as first sheet: {workbook.sheetnames[0]}")

    if "Au" not in workbook.sheetnames:
        return "fail", ["❌ Missing Au sheet."]
    if "Others" not in workbook.sheetnames:
        return "fail", ["❌ Missing Others sheet."]

    au_sheet = workbook["Au"]
    others_sheet = workbook["Others"]
    headers = [cell.value for cell in au_sheet[1]]
    selected_elements = detected_normalized_elements(headers)
    if not selected_elements:
        return "fail", ["❌ Could not detect normalized Au/Ag/Cu/Hg columns after the Normalized header."]
    messages.append(f"✅ Detected normalized sample elements: {'+'.join(selected_elements)}")

    raw_columns = {}
    for element in ELEMENTS:
        column_number = find_header(headers, f"{element} (Wt%)")
        if column_number is not None:
            raw_columns[element] = column_number
    missing = [element for element in selected_elements if element not in raw_columns]
    if missing:
        return "fail", [f"❌ Missing raw wt% columns for: {', '.join(missing)}"]

    extra_high = []
    for element in ("Cu", "Hg"):
        if element in selected_elements or element not in raw_columns:
            continue
        for row_number in data_row_numbers(au_sheet):
            if numeric_value(au_sheet.cell(row=row_number, column=raw_columns[element]).value) > SAMPLE_TYPE_WARNING_THRESHOLD_WT:
                extra_high.append(element)
                break
    if extra_high:
        status = "fail"
        messages.append(
            f"❌ Possible wrong sample type: {', '.join(extra_high)} exceeds {SAMPLE_TYPE_WARNING_THRESHOLD_WT:g} wt%."
        )

    low_normalized_rows = []
    normalized_percentages_by_element = {element: [] for element in selected_elements}
    for row_number in data_row_numbers(au_sheet):
        raw_values = [numeric_value(au_sheet.cell(row=row_number, column=raw_columns[element]).value) for element in selected_elements]
        normalized_total = sum(raw_values)
        if normalized_total < MIN_NORMALIZED_TOTAL_WT:
            low_normalized_rows.append(row_number)
        if normalized_total > 0:
            for element, raw_value in zip(selected_elements, raw_values):
                normalized_percentages_by_element[element].append(raw_value * 100 / normalized_total)

    if low_normalized_rows:
        status = "fail"
        messages.append(
            "❌ Normalized total below "
            f"{MIN_NORMALIZED_TOTAL_WT:g} wt% in Au rows: {', '.join(map(str, low_normalized_rows[:10]))}"
        )
    else:
        messages.append(f"✅ All Au data rows have Normalized total >= {MIN_NORMALIZED_TOTAL_WT:g} wt%.")

    average_sum = 0.0
    for values in normalized_percentages_by_element.values():
        if values:
            average_sum += sum(values) / len(values)
    if abs(average_sum - 100.0) > 1e-9:
        status = "fail"
        messages.append(f"❌ Average normalized chemistry sums to {average_sum:.6f}, not 100.")
    else:
        messages.append("✅ Average normalized chemistry sums to 100.")

    au_count = len(data_row_numbers(au_sheet))
    others_count = max(0, others_sheet.max_row - 1)
    if au_count != others_count:
        if status == "pass":
            status = "warn"
        messages.append(f"⚠️ Others Area rows ({others_count}) do not match Au rows ({au_count}).")
    else:
        messages.append(f"✅ Others Area rows match Au rows ({au_count}).")

    return status, messages


def validate_word_report(word_path: Path) -> tuple[str, list[str]]:
    """Check basic Word package health and embedded Excel object presence."""
    messages = []
    try:
        with ZipFile(word_path) as archive:
            names = archive.namelist()
    except BadZipFile:
        return "fail", ["❌ Selected Word file is not a valid .docx package."]

    embeddings = [name for name in names if name.startswith("word/embeddings/")]
    if not embeddings:
        return "warn", ["⚠️ No embedded Excel worksheet objects were found in the Word file."]
    messages.append(f"✅ Word file contains {len(embeddings)} embedded object(s).")
    return "pass", messages


def combined_status(statuses: list[str]) -> str:
    """Return the highest-severity status."""
    if "fail" in statuses:
        return "fail"
    if "warn" in statuses:
        return "warn"
    return "pass"


def status_icon(status: str) -> str:
    return {"pass": "✅ PASS", "warn": "⚠️ WARNING", "fail": "❌ FAIL"}[status]


def main() -> None:
    root = Tk()
    root.withdraw()

    word_file = filedialog.askopenfilename(
        title="Select final Word report",
        filetypes=(("Word documents", "*.docx"), ("All files", "*.*")),
    )
    if not word_file:
        return
    workbook_file = filedialog.askopenfilename(
        title="Select data workbook with Raw/Au/Others sheets",
        filetypes=(("Excel workbooks", "*.xlsx *.xlsm"), ("All files", "*.*")),
    )
    if not workbook_file:
        return

    word_path = Path(word_file)
    workbook_path = Path(workbook_file)
    data_status, data_messages = validate_data_workbook(workbook_path)
    word_status, word_messages = validate_word_report(word_path)
    final_status = combined_status([data_status, word_status])

    report_lines = [
        f"Au Existing Output QC: {status_icon(final_status)}",
        f"Word file: {word_path}",
        f"Data workbook: {workbook_path}",
        "",
        "Data workbook checks:",
        *data_messages,
        "",
        "Word report checks:",
        *word_messages,
        "",
    ]
    report_path = word_path.with_name(f"{word_path.stem}_QC_Report.txt")
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    messagebox.showinfo(
        "Au existing output QC",
        f"{status_icon(final_status)}\n\nQC report saved:\n{report_path}",
    )


if __name__ == "__main__":
    main()
