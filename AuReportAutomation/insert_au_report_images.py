"""Create the first Au report workbook by inserting reflected light and SEM images.

This script is intentionally focused on step 1 of the Au report automation:
- ask for the sample type
- ask for reflected light and SEM image folders
- sort .jpg/.jpeg images by the first number in each filename
- create a new Excel workbook
- add the correct headers for the selected sample type
- insert reflected light and SEM images into the correct columns
- number rows automatically in column A

Run from PyCharm or a terminal with the project virtual environment active.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from tkinter import Tk, filedialog, messagebox, simpledialog

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage

ROW_HEIGHT = 45
COLUMN_WIDTH = 8.43
IMAGE_EXTENSIONS = {".jpg", ".jpeg"}
# Approximate display size for Excel's default 96 DPI rendering.
TARGET_IMAGE_WIDTH_PX = 64
TARGET_IMAGE_HEIGHT_PX = 60
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


@dataclass(frozen=True)
class SampleLayout:
    """Column layout for one Au report sample type."""

    sample_type: str
    headers: tuple[str, ...]
    reflected_light_column: int
    sem_column: int


SAMPLE_LAYOUTS: dict[str, SampleLayout] = {
    "1": SampleLayout(
        sample_type="Au+Ag",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=4,
        sem_column=5,
    ),
    "2": SampleLayout(
        sample_type="Au+Ag+Cu",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Cu (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=5,
        sem_column=6,
    ),
    "3": SampleLayout(
        sample_type="Au+Ag+Cu+Hg",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Cu (Wt%)", "Hg (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=6,
        sem_column=7,
    ),
    "4": SampleLayout(
        sample_type="Au+Ag+Hg",
        headers=("No.", "Au (Wt%)", "Ag (Wt%)", "Hg (Wt%)", "R. Light Images", "SEM Images"),
        reflected_light_column=5,
        sem_column=6,
    ),
}


class UserCancelledError(Exception):
    """Raised when the user cancels one of the required GUI prompts."""


def get_first_number(path: Path) -> int:
    """Return the first number found in a filename for numeric image sorting."""
    match = re.search(r"\d+", path.stem)
    if not match:
        return 10**12
    return int(match.group())


def sorted_image_files(folder: Path) -> list[Path]:
    """Return .jpg/.jpeg images sorted by their first filename number, then name."""
    image_files = [path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS]
    return sorted(image_files, key=lambda path: (get_first_number(path), path.name.lower()))


def choose_sample_layout() -> SampleLayout:
    """Ask the user which Au sample layout to use."""
    prompt = (
        "Select sample type by number:\n\n"
        "1 = Au+Ag\n"
        "2 = Au+Ag+Cu\n"
        "3 = Au+Ag+Cu+Hg\n"
        "4 = Au+Ag+Hg"
    )
    choice = simpledialog.askstring("Au Report Sample Type", prompt)
    if choice is None:
        raise UserCancelledError("Sample type selection was cancelled.")
    choice = choice.strip()
    if choice not in SAMPLE_LAYOUTS:
        messagebox.showerror("Invalid sample type", "Please run again and enter 1, 2, 3, or 4.")
        raise UserCancelledError(f"Invalid sample type choice: {choice}")
    return SAMPLE_LAYOUTS[choice]


def choose_folder(title: str) -> Path:
    """Ask the user to select a folder."""
    selected = filedialog.askdirectory(title=title)
    if not selected:
        raise UserCancelledError(f"Folder selection was cancelled: {title}")
    return Path(selected)


def choose_output_file(default_name: str) -> Path:
    """Ask the user where to save the new final workbook."""
    selected = filedialog.asksaveasfilename(
        title="Save new Au report workbook as",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=(("Excel workbook", "*.xlsx"),),
    )
    if not selected:
        raise UserCancelledError("Output workbook selection was cancelled.")
    return Path(selected)


def set_standard_dimensions(worksheet, max_row: int, max_column: int) -> None:
    """Apply the requested row height and column width."""
    for row_number in range(1, max_row + 1):
        worksheet.row_dimensions[row_number].height = ROW_HEIGHT
    for column_number in range(1, max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = COLUMN_WIDTH


def fit_image_to_cell(image_path: Path) -> tuple[int, int]:
    """Calculate image dimensions that preserve aspect ratio inside one report cell."""
    with PillowImage.open(image_path) as image:
        original_width, original_height = image.size
    scale = min(TARGET_IMAGE_WIDTH_PX / original_width, TARGET_IMAGE_HEIGHT_PX / original_height)
    return max(1, int(original_width * scale)), max(1, int(original_height * scale))


def add_image_to_cell(worksheet, image_path: Path, row_number: int, column_number: int) -> None:
    """Insert one image into a worksheet cell."""
    image = ExcelImage(str(image_path))
    image.width, image.height = fit_image_to_cell(image_path)
    worksheet.add_image(image, f"{get_column_letter(column_number)}{row_number}")


def prepare_headers(worksheet, layout: SampleLayout) -> None:
    """Add headers for the selected sample layout."""
    for column_number, header in enumerate(layout.headers, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_image_workbook(layout: SampleLayout, reflected_images: list[Path], sem_images: list[Path]) -> Workbook:
    """Create the first Au report workbook with headers, numbers, and images."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = layout.sample_type.replace("+", "_")

    image_count = max(len(reflected_images), len(sem_images))
    max_column = len(layout.headers)
    set_standard_dimensions(worksheet, max_row=image_count + 1, max_column=max_column)
    prepare_headers(worksheet, layout)

    for index in range(image_count):
        row_number = index + 2
        number_cell = worksheet.cell(row=row_number, column=1, value=index + 1)
        number_cell.alignment = Alignment(horizontal="center", vertical="center")

        if index < len(reflected_images):
            add_image_to_cell(worksheet, reflected_images[index], row_number, layout.reflected_light_column)
        if index < len(sem_images):
            add_image_to_cell(worksheet, sem_images[index], row_number, layout.sem_column)

    worksheet.freeze_panes = "A2"
    return workbook


def main() -> None:
    """Run the GUI workflow."""
    root = Tk()
    root.withdraw()

    layout = choose_sample_layout()
    reflected_folder = choose_folder("Select reflected light image folder")
    sem_folder = choose_folder("Select SEM image folder")

    reflected_images = sorted_image_files(reflected_folder)
    sem_images = sorted_image_files(sem_folder)
    if not reflected_images and not sem_images:
        messagebox.showerror("No images found", "No .jpg or .jpeg images were found in either selected folder.")
        raise UserCancelledError("No images found in selected folders.")

    output_file = choose_output_file(f"Au_Report_{layout.sample_type.replace('+', '_')}.xlsx")
    workbook = create_image_workbook(layout, reflected_images, sem_images)
    workbook.save(output_file)

    messagebox.showinfo(
        "Au report workbook created",
        "Workbook saved successfully.\n\n"
        f"Output: {output_file}\n"
        f"Reflected light images inserted: {len(reflected_images)}\n"
        f"SEM images inserted: {len(sem_images)}",
    )


if __name__ == "__main__":
    main()
