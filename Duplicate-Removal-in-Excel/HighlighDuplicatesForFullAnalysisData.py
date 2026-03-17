from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import (
    Tk, filedialog, Listbox, MULTIPLE, Button, Label, Scrollbar,
    END, RIGHT, Y, LEFT, BOTH, Frame
)
from tkinter.simpledialog import askfloat
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ===========================
# SETTINGS
# ===========================
DEFAULT_TOL_UM = 20.0               # you can set your preferred default here
REQUIRE_DIFFERENT_FIELD = True      # keep True for overlap duplicates
PROGRESS_EVERY_N_ROWS = 5000

BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

# Group colors (cycle). Light/pastel fills so text stays readable.
GROUP_COLORS = [
    "FFF59D",  # yellow
    "B3E5FC",  # light blue
    "C8E6C9",  # light green
    "FFCCBC",  # light orange
    "E1BEE7",  # light purple
    "F8BBD0",  # light pink
    "D7CCC8",  # light brown/grey
    "DCEDC8",  # pale green
    "BBDEFB",  # pale blue
    "FFE0B2",  # pale orange
]


# ===========================
# UI helpers
# ===========================
def pick_file() -> Path:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    p = filedialog.askopenfilename(
        title="Select the 'Classified xxxxxx' Excel file",
        filetypes=[("Excel files", "*.xlsx")],
    )
    root.destroy()
    if not p:
        raise SystemExit("No file selected.")
    return Path(p)


def select_sheets_gui(sheetnames: List[str]) -> List[str]:
    root = Tk()
    root.title("Select sheets to process (duplicates only)")
    root.attributes("-topmost", True)

    Label(root, text="Select sheets (Ctrl/Shift-click).").pack(padx=10, pady=8)

    frame = Frame(root)
    frame.pack(padx=10, pady=5, fill=BOTH, expand=True)

    scrollbar = Scrollbar(frame)
    scrollbar.pack(side=RIGHT, fill=Y)

    lb = Listbox(frame, selectmode=MULTIPLE, yscrollcommand=scrollbar.set, width=60, height=18)
    for name in sheetnames:
        lb.insert(END, name)
    lb.pack(side=LEFT, fill=BOTH, expand=True)
    scrollbar.config(command=lb.yview)

    selected: List[str] = []

    def select_all():
        lb.select_set(0, END)

    def ok():
        for i in lb.curselection():
            selected.append(lb.get(i))
        root.destroy()

    btn_frame = Frame(root)
    btn_frame.pack(pady=10)

    Button(btn_frame, text="Select All", command=select_all, width=12).pack(side=LEFT, padx=6)
    Button(btn_frame, text="OK", command=ok, width=12).pack(side=LEFT, padx=6)

    root.mainloop()
    return selected


# ===========================
# Core helpers
# ===========================
def norm_header(s: str) -> str:
    return (s or "").strip().lower().replace("\n", " ").replace("µ", "u")


def find_col_idx(headers: List[str], candidates: List[str]) -> Optional[int]:
    nheaders = [norm_header(h) for h in headers]
    for c in candidates:
        c = c.lower()
        for i, h in enumerate(nheaders, start=1):
            if c in h:
                return i
    return None


def is_um_header(h: str) -> bool:
    return "um" in norm_header(h)


def safe_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
    try:
        return float(v)
    except Exception:
        return None


def safe_str(v) -> str:
    return "" if v is None else str(v).strip()


@dataclass
class Pt:
    excel_row: int
    feature_id: str
    field_id: str
    x_um: float
    y_um: float


class UnionFind:
    def __init__(self, n: int):
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, a: int) -> int:
        while self.parent[a] != a:
            self.parent[a] = self.parent[self.parent[a]]
            a = self.parent[a]
        return a

    def union(self, a: int, b: int) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
        elif self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
        else:
            self.parent[rb] = ra
            self.rank[ra] += 1


def grid_key(x_um: float, y_um: float, cell: float) -> Tuple[int, int]:
    return (int(math.floor(x_um / cell)), int(math.floor(y_um / cell)))


def build_duplicate_groups(ws, tol_um: float) -> Tuple[Optional[str], List[str], Dict[int, List[int]], List[Pt], List[str]]:
    """
    Returns: (error_note, headers, dup_groups_map, pts, headers_list)
    dup_groups_map: group_id -> list of point indices
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2:
        return ("Empty sheet", [], {}, [], [])

    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    headers_str = ["" if h is None else str(h) for h in headers]

    col_x = find_col_idx(headers_str, ["stage x"])
    col_y = find_col_idx(headers_str, ["stage y"])
    col_feat = find_col_idx(headers_str, ["feature id", "feature"])
    col_field = find_col_idx(headers_str, ["field"])

    if not (col_x and col_y and col_feat and col_field):
        return ("Missing one of required columns (Stage X, Stage Y, Feature, Field)", headers_str, {}, [], headers_str)

    x_hdr = headers_str[col_x - 1]
    y_hdr = headers_str[col_y - 1]
    scale = 1.0 if (is_um_header(x_hdr) or is_um_header(y_hdr)) else 1000.0  # mm -> µm

    pts: List[Pt] = []
    for r in range(2, max_row + 1):
        x = safe_float(ws.cell(row=r, column=col_x).value)
        y = safe_float(ws.cell(row=r, column=col_y).value)
        if x is None or y is None:
            continue
        pts.append(
            Pt(
                excel_row=r,
                feature_id=safe_str(ws.cell(row=r, column=col_feat).value),
                field_id=safe_str(ws.cell(row=r, column=col_field).value),
                x_um=x * scale,
                y_um=y * scale,
            )
        )

    n = len(pts)
    if n < 2:
        return ("Not enough points", headers_str, {}, pts, headers_str)

    # bucket points
    buckets: Dict[Tuple[int, int], List[int]] = {}
    for i, p in enumerate(pts):
        buckets.setdefault(grid_key(p.x_um, p.y_um, tol_um), []).append(i)

    uf = UnionFind(n)

    # find duplicates
    for i, p in enumerate(pts):
        kx, ky = grid_key(p.x_um, p.y_um, tol_um)
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for j in buckets.get((kx + dx, ky + dy), []):
                    if j <= i:
                        continue
                    q = pts[j]
                    if REQUIRE_DIFFERENT_FIELD and p.field_id == q.field_id:
                        continue
                    dist = math.hypot(p.x_um - q.x_um, p.y_um - q.y_um)
                    if dist <= tol_um:
                        uf.union(i, j)

    groups: Dict[int, List[int]] = {}
    for i in range(n):
        groups.setdefault(uf.find(i), []).append(i)

    # Keep only groups > 1
    dup_groups = [g for g in groups.values() if len(g) > 1]
    if not dup_groups:
        return (None, headers_str, {}, pts, headers_str)

    # Assign group IDs 1..N
    dup_groups_map: Dict[int, List[int]] = {}
    gid = 1
    for g in dup_groups:
        dup_groups_map[gid] = g
        gid += 1

    return (None, headers_str, dup_groups_map, pts, headers_str)


def copy_row(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        dst_ws.cell(row=dst_row, column=c, value=src_ws.cell(row=src_row, column=c).value)


# ===========================
# Main
# ===========================
def main():
    in_path = pick_file()
    print(f"Selected: {in_path}", flush=True)

    print("Loading workbook...", flush=True)
    wb = load_workbook(in_path)
    print(f"Loaded. Sheets: {len(wb.sheetnames)}", flush=True)

    chosen = select_sheets_gui(wb.sheetnames)
    if not chosen:
        raise SystemExit("No sheets selected.")
    print(f"Selected sheets ({len(chosen)}): {chosen}", flush=True)

    # Tolerance dialog (topmost)
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    tol = askfloat(
        "Duplicate tolerance (µm)",
        "Enter XY tolerance (µm). Example: 20, 25 ...",
        initialvalue=DEFAULT_TOL_UM,
        minvalue=0.5,
        maxvalue=500.0,
        parent=root,
    )
    root.destroy()
    if tol is None:
        raise SystemExit("Cancelled.")
    print(f"Tolerance = {tol} µm | Require different field = {REQUIRE_DIFFERENT_FIELD}", flush=True)

    out_wb = Workbook()
    # remove default blank sheet
    out_wb.remove(out_wb.active)

    report_rows = []

    for sheet_name in chosen:
        src_ws = wb[sheet_name]
        print(f"\nProcessing: {sheet_name}", flush=True)

        note, headers, dup_groups_map, pts, headers_list = build_duplicate_groups(src_ws, tol_um=tol)

        max_col = src_ws.max_column
        if note and note != "Not enough points":
            print(f"  NOTE: {note}", flush=True)

        if not dup_groups_map:
            report_rows.append({
                "Sheet": sheet_name,
                "PointsChecked": len(pts),
                "DuplicateRows": 0,
                "DuplicateGroups": 0,
                "Note": note or ""
            })
            # Still create an empty sheet with just header so you know it was processed
            dst_ws = out_wb.create_sheet(sheet_name[:31])
            # copy header
            for c in range(1, max_col + 1):
                dst_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)
            continue

        # Create output sheet and copy header
        dst_ws = out_wb.create_sheet(sheet_name[:31])
        for c in range(1, max_col + 1):
            dst_ws.cell(row=1, column=c, value=src_ws.cell(row=1, column=c).value)

        # Build a mapping: src_row -> group_id
        srcrow_to_gid: Dict[int, int] = {}
        for gid, idxs in dup_groups_map.items():
            for idx in idxs:
                srcrow_to_gid[pts[idx].excel_row] = gid

        # Copy ONLY duplicate rows, maintaining original order
        dst_row = 2
        for src_row in range(2, src_ws.max_row + 1):
            if src_row in srcrow_to_gid:
                copy_row(src_ws, dst_ws, src_row, dst_row, max_col)

                gid = srcrow_to_gid[src_row]
                color = GROUP_COLORS[(gid - 1) % len(GROUP_COLORS)]
                fill = PatternFill("solid", fgColor=color)

                # Fill entire row
                for c in range(1, max_col + 1):
                    dst_ws.cell(row=dst_row, column=c).fill = fill
                dst_row += 1

        dup_rows = dst_row - 2
        dup_groups = len(dup_groups_map)

        report_rows.append({
            "Sheet": sheet_name,
            "PointsChecked": len(pts),
            "DuplicateRows": dup_rows,
            "DuplicateGroups": dup_groups,
            "Note": note or ""
        })

        print(f"  Duplicate groups: {dup_groups} | Rows kept: {dup_rows}", flush=True)

    # Add report sheet
    rep = out_wb.create_sheet("Duplicate_Report", 0)
    rep_headers = ["Sheet", "PointsChecked", "DuplicateRows", "DuplicateGroups", "Note"]
    for c, h in enumerate(rep_headers, start=1):
        cell = rep.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT

    r = 2
    for row in report_rows:
        rep.cell(row=r, column=1, value=row["Sheet"])
        rep.cell(row=r, column=2, value=row["PointsChecked"])
        rep.cell(row=r, column=3, value=row["DuplicateRows"])
        rep.cell(row=r, column=4, value=row["DuplicateGroups"])
        rep.cell(row=r, column=5, value=row["Note"])
        r += 1

    rep.cell(row=r + 1, column=1, value="Tolerance (µm)")
    rep.cell(row=r + 1, column=2, value=tol)
    rep.cell(row=r + 2, column=1, value="Require different field?")
    rep.cell(row=r + 2, column=2, value=str(REQUIRE_DIFFERENT_FIELD))

    out_path = in_path.with_name(in_path.stem + "_DUPS_ONLY.xlsx")
    print(f"\nSaving: {out_path}", flush=True)
    out_wb.save(out_path)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()