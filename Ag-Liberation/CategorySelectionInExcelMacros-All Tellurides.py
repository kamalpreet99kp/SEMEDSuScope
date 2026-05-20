import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
#  NEW HEADERS
# ============================================================

# Summary column + your class columns
CLASS_HEADERS = [
    "F",
    "Te-A", "Te-E",
    "Py-A", "Py-E",
    "CGS-A", "CGS-E",
    "Gg-A", "Gg-E",
    "AuAg-A", "AuAg-E",
]
NEW_HEADERS = ["Summary"] + CLASS_HEADERS  # inserted after Image

HEADER_ROW = 1
CLICK_MARKER = "x"
CLEAR_OTHER_CLASS_CELLS = True

# ============================================================
#  COLORS
#  - paired categories use same fill color
#  - Te-A / Te-E
#  - Py-A / Py-E
#  - CGS-A / CGS-E
#  - Gg-A / Gg-E
#  - AuAg-A / AuAg-E
#  - Summary and F have fixed colors
# ============================================================

# Choose easy-to-see pastel groups + distinct singles
COLOR_SUMMARY = "FFF200"   # yellow
COLOR_F       = "FF4D4D"   # red

COLOR_TE   = "F8CBAD"      # Te pair (peach)
COLOR_PY   = "D9E1F2"      # Py pair (light blue)
COLOR_CGS  = "E4DFEC"      # CGS pair (light purple)
COLOR_GG   = "C6EFCE"      # Gg pair (light green)
COLOR_AUAG = "FFE699"      # AuAg pair (light gold)

COLORS = {"Summary": COLOR_SUMMARY, "F": COLOR_F}
for h in ("Te-A", "Te-E"):
    COLORS[h] = COLOR_TE
for h in ("Py-A", "Py-E"):
    COLORS[h] = COLOR_PY
for h in ("CGS-A", "CGS-E"):
    COLORS[h] = COLOR_CGS
for h in ("Gg-A", "Gg-E"):
    COLORS[h] = COLOR_GG
for h in ("AuAg-A", "AuAg-E"):
    COLORS[h] = COLOR_AUAG

# Column width: ~8 characters
NEWCOL_WIDTH = 9  # Excel uses approx character units


# -----------------------------
# Live logger window
# -----------------------------
class LiveLogger:
    def __init__(self, title="Progress"):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.geometry("900x520")
        self.root.attributes("-topmost", True)

        self.text = tk.Text(self.root, wrap="word")
        self.text.pack(fill="both", expand=True)

        self.text.insert("end", "=== Live Output ===\n")
        self.text.see("end")
        self.root.update()

    def log(self, msg: str):
        print(msg, flush=True)
        self.text.insert("end", msg + "\n")
        self.text.see("end")
        self.root.update()


# -----------------------------
# GUI: sheet selection
# -----------------------------
def choose_sheets_gui(sheet_names):
    win = tk.Toplevel()
    win.title("Select sheets to update (All Tellurides)")
    win.geometry("520x600")

    tk.Label(
        win,
        text="Select sheets to update.\nCtrl/Shift multi-select.\n"
             "If it looks stuck, this window might be behind other windows.",
        justify="left",
        anchor="w"
    ).pack(fill="x", padx=10, pady=10)

    frame = tk.Frame(win)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")

    lb = tk.Listbox(frame, selectmode="extended", yscrollcommand=scrollbar.set)
    for s in sheet_names:
        lb.insert("end", s)
    lb.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=lb.yview)

    default_skip = {"Area", "Check Report", "Summary", "Raw Data"}
    for i, s in enumerate(sheet_names):
        if s not in default_skip:
            lb.selection_set(i)

    result = {"selected": None}

    def ok():
        idxs = lb.curselection()
        result["selected"] = [sheet_names[i] for i in idxs]
        win.destroy()

    def cancel():
        result["selected"] = None
        win.destroy()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=(0, 10))
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.grab_set()
    win.wait_window()
    return result["selected"]


# -----------------------------
# Helpers
# -----------------------------
def find_header_col(ws, header_text):
    for cell in ws[HEADER_ROW]:
        if str(cell.value).strip() == header_text:
            return cell.col_idx
    return None


def freeze_top_row(ws):
    ws.freeze_panes = "A2"


def make_border():
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_fill_align_border_to_range(ws, col_idx, start_row, end_row, rgb):
    fill = PatternFill("solid", fgColor=rgb)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = make_border()
    for r in range(start_row, end_row + 1):
        cell = ws.cell(r, col_idx)
        cell.fill = fill
        cell.alignment = align
        cell.border = border


# -----------------------------
# Insert new columns after Image
# -----------------------------
def add_columns_after_image(ws):
    img_col = find_header_col(ws, "Image")
    if img_col is None:
        raise ValueError(f"Sheet '{ws.title}': Could not find 'Image' header in row 1.")

    first_new_col = img_col + 1
    ws.insert_cols(first_new_col, amount=len(NEW_HEADERS))

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = make_border()

    max_row = ws.max_row

    for i, h in enumerate(NEW_HEADERS):
        c = first_new_col + i

        # header cell
        hc = ws.cell(HEADER_ROW, c, h)
        hc.font = header_font
        hc.alignment = header_align
        hc.border = header_border
        hc.fill = PatternFill("solid", fgColor=COLORS.get(h, "FFFFFF"))

        # whole column formatting (rows 2..end)
        apply_fill_align_border_to_range(ws, c, 2, max_row, COLORS.get(h, "FFFFFF"))

        # narrow width
        ws.column_dimensions[get_column_letter(c)].width = NEWCOL_WIDTH if h != "Summary" else 12

    ws.row_dimensions[1].height = 20
    return img_col, first_new_col


# -----------------------------
# Counts block (to the right)
# -----------------------------
def add_counts_block(ws, first_new_col):
    """
    Adds a small counts table to the right of the new columns.
    Counts are based on Summary column values.
    """
    summary_col = first_new_col  # Summary is first new column
    summary_letter = get_column_letter(summary_col)

    start_col = first_new_col + len(NEW_HEADERS) + 1  # after new cols + 1 blank
    start_row = 1

    bold = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center")
    border = make_border()

    ws.cell(start_row, start_col, "Counts").font = bold
    ws.cell(start_row, start_col).alignment = align
    ws.cell(start_row, start_col).border = border

    ws.cell(start_row, start_col + 1, "Total").font = bold
    ws.cell(start_row, start_col + 1).alignment = align
    ws.cell(start_row, start_col + 1).border = border

    summary_range = f"{summary_letter}:{summary_letter}"

    for i, cls in enumerate(CLASS_HEADERS, start=1):
        a = ws.cell(start_row + i, start_col, cls)
        a.font = bold
        a.alignment = align
        a.border = border

        b = ws.cell(start_row + i, start_col + 1, f'=COUNTIF({summary_range},"{cls}")')
        b.alignment = align
        b.border = border

    ws.column_dimensions[get_column_letter(start_col)].width = 14
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 10


# -----------------------------
# VBA macro (auto-updated for new class count)
# -----------------------------
# lastClassCol = imgCol + (1 Summary + len(CLASS_HEADERS)) = imgCol + len(NEW_HEADERS)
VBA_CODE = r'''
Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    On Error GoTo SafeExit

    If Target Is Nothing Then GoTo SafeExit
    If Target.CountLarge <> 1 Then GoTo SafeExit
    If Target.Row < 2 Then GoTo SafeExit

    Dim imgCol As Long
    imgCol = 0

    Dim c As Long
    For c = 1 To 800
        If Trim(CStr(Me.Cells(1, c).Value)) = "Image" Then
            imgCol = c
            Exit For
        End If
    Next c

    If imgCol = 0 Then GoTo SafeExit

    Dim summaryCol As Long
    summaryCol = imgCol + 1
    Dim firstClassCol As Long
    firstClassCol = imgCol + 2
    Dim lastClassCol As Long
    lastClassCol = imgCol + ''' + str(len(NEW_HEADERS)) + r'''

    If Target.Column < firstClassCol Or Target.Column > lastClassCol Then
        GoTo SafeExit
    End If

    Application.EnableEvents = False
    Me.Cells(Target.Row, summaryCol).Value = Me.Cells(1, Target.Column).Value

''' + ("    If True Then\n" if CLEAR_OTHER_CLASS_CELLS else "    If False Then\n") + r'''
        Dim cc As Long
        For cc = firstClassCol To lastClassCol
            If cc <> Target.Column Then
                Me.Cells(Target.Row, cc).Value = ""
            End If
        Next cc
    End If

    Me.Cells(Target.Row, Target.Column).Value = "''' + CLICK_MARKER + r'''"

SafeExit:
    Application.EnableEvents = True
End Sub
'''


def inject_vba_to_sheets_xlsm(xlsx_path: Path, out_xlsm_path: Path, target_sheets: list, logger):
    """
    Robust VBA injection using VBProject.VBComponents(CodeName).CodeModule
    """
    import win32com.client as win32

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = True

    try:
        try:
            excel.AutomationSecurity = 1  # msoAutomationSecurityLow
        except Exception:
            pass

        logger.log(f"➡️ [Excel] Opening workbook: {xlsx_path}")
        wb = excel.Workbooks.Open(str(xlsx_path), ReadOnly=False)

        logger.log(f"➡️ [Excel] Saving as XLSM: {out_xlsm_path}")
        wb.SaveAs(str(out_xlsm_path), FileFormat=52)

        vbproj = wb.VBProject

        comp_names = [vbproj.VBComponents(i).Name for i in range(1, vbproj.VBComponents.Count + 1)]
        logger.log(f"✅ [Excel] VBComponents visible: {len(comp_names)}")

        for sname in target_sheets:
            logger.log(f"➡️ [Excel] Injecting macro into sheet: {sname}")
            ws = wb.Worksheets(sname)
            codename = ws.CodeName
            logger.log(f"   - CodeName: {codename}")

            vbcomp = vbproj.VBComponents(codename)
            cm = vbcomp.CodeModule

            if cm.CountOfLines > 0:
                cm.DeleteLines(1, cm.CountOfLines)
            cm.AddFromString(VBA_CODE)

            logger.log(f"✅ VBA injected into: {sname}")

        logger.log("➡️ [Excel] Saving workbook...")
        wb.Save()
        wb.Close(SaveChanges=True)
        logger.log("✅ [Excel] Macro injection complete.")

    finally:
        try:
            excel.Quit()
        except Exception:
            pass


def main():
    logger = LiveLogger("Click-Columns Builder - Live Output")
    root = tk.Tk()
    root.withdraw()

    logger.log("Select the MASTER Excel file now...")
    master_path = filedialog.askopenfilename(
        title="Select MASTER Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    if not master_path:
        logger.log("No file selected. Exiting.")
        return

    master_path = Path(master_path)
    logger.log(f"Selected: {master_path}")

    out_xlsx = master_path.with_name(master_path.stem + "_with_click_columns.xlsx")
    out_xlsm = master_path.with_name(master_path.stem + "_with_click_columns.xlsm")

    # quick lock check
    for p in [out_xlsx, out_xlsm]:
        if p.exists():
            try:
                p.rename(p)
            except Exception:
                messagebox.showerror("File is open/locked", f"Close this file in Excel first:\n{p}")
                return

    logger.log("Loading workbook (openpyxl)...")
    wb = load_workbook(master_path)
    all_sheets = wb.sheetnames
    logger.log(f"Workbook loaded. Sheets: {len(all_sheets)}")

    logger.log("Select sheets window is opening (it might appear behind other windows).")
    selected_sheets = choose_sheets_gui(all_sheets)
    if not selected_sheets:
        logger.log("No sheets selected. Exiting.")
        return

    logger.log(f"Sheets selected: {len(selected_sheets)}")

    updated, skipped = [], []

    for i, sname in enumerate(selected_sheets, start=1):
        logger.log(f"[{i}/{len(selected_sheets)}] Processing sheet: {sname}")
        ws = wb[sname]

        img_col = find_header_col(ws, "Image")
        if img_col is None:
            logger.log(f"  ⚠️ Skipped (no 'Image' header): {sname}")
            skipped.append(sname)
            continue

        freeze_top_row(ws)
        logger.log("  ✅ Freeze top row set (A2).")

        _, first_new_col = add_columns_after_image(ws)
        logger.log(f"  ✅ Added {len(NEW_HEADERS)} columns after Image.")

        add_counts_block(ws, first_new_col)
        logger.log("  ✅ Added counts block.")

        updated.append(sname)

    logger.log(f"Saving intermediate XLSX: {out_xlsx}")
    wb.save(out_xlsx)
    logger.log("✅ Intermediate XLSX saved.")

    try:
        logger.log("Starting Excel macro injection step...")
        inject_vba_to_sheets_xlsm(out_xlsx, out_xlsm, updated, logger)
        messagebox.showinfo(
            "Success",
            f"Created:\n{out_xlsm}\n\nUpdated sheets: {len(updated)}\nSkipped: {len(skipped)}\n\n"
            f"Open .xlsm and enable macros."
        )
        logger.log("DONE ✅")
    except Exception as e:
        logger.log(f"❌ Macro injection failed: {e}")
        messagebox.showwarning(
            "Partial Success",
            f"Columns/freeze/counts saved as:\n{out_xlsx}\n\nMacro injection failed:\n{e}"
        )


if __name__ == "__main__":
    main()