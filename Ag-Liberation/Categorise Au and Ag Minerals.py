import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


# ============================================================
#  AU + AG MINERAL CLASSIFICATION
#  - Au minerals are evaluated first, then Ag minerals
#  - Uses practical EDS wt% windows based on stoichiometry + Ag script style
#  - Applies "association caps" for non-member elements to reduce confusion
#  - Writes mineral sheets only when rows exist
# ============================================================


def v(row, col):
    x = row.get(col, 0)
    return 0 if pd.isna(x) else x


TRACE = 2.0
MINOR = 6.0
AUAG_CUTOFF = 2.0


def classify_mineral(row):
    ag = v(row, "Ag (Wt%)")
    s = v(row, "S (Wt%)")
    hg = v(row, "Hg (Wt%)")
    te = v(row, "Te (Wt%)")
    se = v(row, "Se (Wt%)")
    sb = v(row, "Sb (Wt%)")
    ars = v(row, "As (Wt%)")
    fe = v(row, "Fe (Wt%)")
    cu = v(row, "Cu (Wt%)")
    bi = v(row, "Bi (Wt%)")
    au = v(row, "Au (Wt%)")
    o = v(row, "O (Wt%)")

    # Global gate for this Au+Ag liberation workflow:
    # rows with both Ag and Au below cutoff are not classified.
    if ag <= AUAG_CUTOFF and au <= AUAG_CUTOFF:
        return None

    # ========================================================
    # AU MINERALS FIRST
    # ========================================================

    # 1) Native Au
    if (30 <= au <= 100 and ag < 1 and te < 2 and bi < 2 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Native Au"

    # 2) Au-Ag (Electrum)
    if (5 <= au <= 98 and 5 <= ag <= 98 and te < 2 and bi < 2 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Au-Ag (Electrum)"

    # 3) Maldonite (Au2Bi)
    if (35 <= au <= 70 and 18 <= bi <= 70 and ag < 6 and te < 6 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Maldonite"

    # 4) Calaverite (AuTe2)
    if (35 <= au <= 50 and 45 <= te <= 63 and ag < 6 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE and bi < 6):
        return "Calaverite"

    # 5) Sylvanite ((Au,Ag)Te2)
    if (17 <= au <= 28 and 6 <= ag <= 15 and 49 <= te <= 65 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE and bi < 6):
        return "Sylvanite"

    # 6) Petzite (Ag3AuTe2)
    if (30 <= ag <= 48 and 27 <= te <= 36 and 17 <= au <= 29 and
            s < TRACE and se < TRACE and bi < 6 and
            hg < TRACE and sb < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE):
        return "Petzite"

    # ========================================================
    # AG MINERALS (existing script logic + ordering)
    # ========================================================

    if (se >= 6 and bi >= 15 and te < TRACE and s < TRACE and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= se <= 85 and
            au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE):
        return "Bohdanowiczite"

    if (te >= 6 and bi >= 15 and se < TRACE and s < TRACE and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= te <= 90 and
            au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE):
        return "Volynskite"

    # 8) Hessite (Ag2Te) + associations
    if (40 <= ag <= 68 and 18 <= te <= 45 and au < 5 and
            s < TRACE and se < TRACE and
            hg < TRACE and sb < TRACE and ars < TRACE and
            cu < MINOR and fe < MINOR and bi < MINOR):
        return "Hessite & Associations"

    if (ag >= 30 and sb >= 6 and
            s < TRACE and te < TRACE and se < TRACE and
            ars < TRACE and hg < TRACE and bi < TRACE and
            au < TRACE and cu < TRACE and fe < TRACE):
        return "Dyscrasite"

    if (ag >= 85 and o <= 20 and
            s < TRACE and te < TRACE and se < TRACE and
            sb < TRACE and ars < TRACE and hg < TRACE and bi < TRACE and
            au < MINOR and cu < TRACE and fe < TRACE):
        return "Native Ag"

    if ag > 10 and s > 7 and hg < 7 and te < 6 and se < 6 and bi < 40 and sb < 8 and ars < 8 and cu < 25 and fe < 26:
        return "Acanthite & Associations"

    if 5 < ag < 75 and 5 < s < 31 and 6.99 < hg < 53 and te < 6 and sb < 15 and ars < 15 and bi < 40 and cu < 25 and fe < 26 and se < 6:
        return "Imiterite & Associations"

    if 5 < ag < 75 and s > 7.0 and hg < 12 and fe < 21 and te < 6 and ars > 2 and bi < 40 and cu < 25 and se < 6:
        return "Sulfosalt (Sb & or As ) & Asso"

    if 5 < ag < 75 and s > 4 and hg < 12 and te < 6 and sb > 6 and ars < 2 and bi < 40 and cu < 25 and fe < 21 and se < 6:
        return "Sulfosalt (Sb) & Asso"

    if ag > AUAG_CUTOFF and s > 15 and hg < 6 and te < 6 and ars > 10 and cu > 17 and bi < 40 and se < 6:
        return "Ag Associations with Enargite"

    if ag > AUAG_CUTOFF and hg > 3 and fe < 20 and te < 6 and bi < 40 and cu < 20 and se < 6:
        return "Other Ag-Hg Associations"

    if ag > 2 and s > 2.5 and hg < 4 and te < 6 and fe < 28 and bi < 40 and cu < 28 and se >= 6:
        return "Aguilarite & Associations"

    if ag > AUAG_CUTOFF and (fe > 10 or cu > 10) and bi < 40:
        return "Ag Associations with Sulphides"

    if ag > AUAG_CUTOFF or au > AUAG_CUTOFF:
        return "All Others"

    return None


tk.Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Select Full Analysis Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    print("No file selected.")
    raise SystemExit

book = pd.read_excel(file_path, sheet_name=None)
raw_sheet_name = list(book.keys())[0]
df = book[raw_sheet_name].copy()

df["Mineral Type"] = df.apply(classify_mineral, axis=1)

area_column = "Area (sq. µm)"

all_categories_in_order = [
    "Native Au",
    "Au-Ag (Electrum)",
    "Maldonite",
    "Calaverite",
    "Sylvanite",
    "Petzite",
    "Native Ag",
    "Hessite & Associations",
    "Acanthite & Associations",
    "Bohdanowiczite",
    "Volynskite",
    "Dyscrasite",
    "Imiterite & Associations",
    "Sulfosalt (Sb & or As ) & Asso",
    "Sulfosalt (Sb) & Asso",
    "Ag Associations with Enargite",
    "Other Ag-Hg Associations",
    "Aguilarite & Associations",
    "Ag Associations with Sulphides",
    "All Others",
]

classified_sheets = {}
summary_rows = []

for cat in all_categories_in_order:
    df_cat = df[df["Mineral Type"] == cat]
    if len(df_cat) == 0:
        continue
    classified_sheets[cat] = df_cat.drop(columns=["Mineral Type"])
    summary_rows.append((cat, df_cat[area_column].sum()))

summary_df = pd.DataFrame(summary_rows, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
total_area = summary_df["Total Sum of Area (sq. µm)"].sum()
summary_df["Percentage"] = (summary_df["Total Sum of Area (sq. µm)"] / total_area * 100) if total_area else 0

output_path = os.path.join(os.path.dirname(file_path), f"Classified_AuAg_{os.path.basename(file_path)}")

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)

    for cat in all_categories_in_order:
        if cat in classified_sheets:
            classified_sheets[cat].to_excel(writer, sheet_name=cat, index=False)

    df_auag = df[(df["Ag (Wt%)"] > AUAG_CUTOFF) | (df["Au (Wt%)"] > AUAG_CUTOFF)]
    raw_auag_count = len(df_auag)
    raw_auag_area = df_auag[area_column].sum()

    classified_concat = pd.concat(classified_sheets.values(), ignore_index=True) if classified_sheets else pd.DataFrame()
    classified_rows = df[df["Mineral Type"].notna()]
    classified_total_count = len(classified_rows)
    classified_total_area = classified_rows[area_column].sum()

    unclassified_auag = df_auag[df_auag["Mineral Type"].isna()]
    if len(unclassified_auag) > 0:
        unclassified_auag.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Unclassified_AuAg", index=False)

    duplicate_metric = "N/A"
    if not classified_concat.empty:
        duplicate_keys = [c for c in ["Feature", "Row"] if c in classified_concat.columns]
        if duplicate_keys:
            duplicate_count = int(classified_concat.duplicated(subset=duplicate_keys, keep=False).sum())
            duplicate_metric = f"✅ No duplicates" if duplicate_count == 0 else f"❌ Duplicates found: {duplicate_count}"

    integrity_data = {
        "Metric": [
            f"Rows with Ag or Au > {AUAG_CUTOFF}% in Raw Data",
            "Total rows in classified sheets (non-empty only)",
            f"Area in Raw Data (Ag or Au > {AUAG_CUTOFF}%)",
            "Area in classified sheets (non-empty only)",
            "Area Match",
            "Row Count Match",
            "Unclassified rows with Ag or Au > cutoff",
            "Feature/Row duplicate check in classified sheets"
        ],
        "Value": [
            raw_auag_count,
            classified_total_count,
            round(raw_auag_area, 4),
            round(classified_total_area, 4),
            "✅ Match" if abs(raw_auag_area - classified_total_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_auag_count == classified_total_count else "❌ Mismatch",
            len(unclassified_auag),
            duplicate_metric
        ]
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)

print(f"\n✅ Classification complete.\nOutput saved to:\n{output_path}")

highlight_colors = {
    "Feature": "00BFCF",
    "Area (sq. µm)": "FFEB3B",
    "S (Wt%)": "FFA07A",
    "Fe (Wt%)": "ADD8E6",
    "Cu (Wt%)": "90EE90",
    "As (Wt%)": "D87093",
    "Ag (Wt%)": "F44336",
    "Sb (Wt%)": "A9A9A9",
    "Hg (Wt%)": "9370DB",
    "Se (Wt%)": "FFA09A",
    "Te (Wt%)": "D87099",
    "Au (Wt%)": "FFF2CC",
    "Bi (Wt%)": "E2EFDA",
    "O (Wt%)": "D9E1F2",
}

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = load_workbook(output_path)

for sheet in wb.sheetnames:
    if sheet in ["Area", "Integrity Check"]:
        continue
    ws = wb[sheet]
    if ws.max_row < 2:
        continue

    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = fill

if "Summary" in wb.sheetnames:
    ws = wb["Summary"]

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for cell in ws["A"]:
        cell.font = Font(bold=True)

wb.save(output_path)
