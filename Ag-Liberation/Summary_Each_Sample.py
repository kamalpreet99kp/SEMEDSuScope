import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IGNORE_SHEETS = {"Area", "Summary", "Raw Data", "Integrity Check", "Check Report"}
REQUIRED_HEADERS = ["Feature", "New Brea", "New Len", "Asso"]
HEADER_BLANK_ROWS = 5
BLOCK_GAP_COLUMNS = 2


class LiveLogger:
    def __init__(self, root: tk.Tk):
        self.win = tk.Toplevel(root)
        self.win.title("Summary Each Sample - Live Progress")
        self.win.geometry("980x540")
        self.text = tk.Text(self.win, wrap="word")
        self.text.pack(fill="both", expand=True)
        self.write("=== Summary Each Sample ===")

    def write(self, msg: str):
        self.text.insert("end", msg + "\n")
        self.text.see("end")
        self.win.update_idletasks()
        self.win.update()


def clean_fraction_name(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"^Complete-MASTER_Classified_Tellurides_", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\s*\(AuTe\s*Scan\)\s*$", "", stem, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", stem).strip() or stem


def choose_sample_folder(root: tk.Tk) -> Optional[Path]:
    folder = filedialog.askdirectory(parent=root, title="Select SAMPLE folder containing size-fraction Excel files")
    if not folder:
        return None
    return Path(folder)


def find_excel_files(sample_dir: Path) -> List[Path]:
    files = [p for p in sample_dir.iterdir() if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}]
    return sorted(files, key=lambda p: p.name.lower())


def choose_file_order_gui(root: tk.Tk, files: Sequence[Path]) -> List[Path]:
    win = tk.Toplevel(root)
    win.title("Select and order size-fraction files")
    win.geometry("980x620")

    tk.Label(
        win,
        text=(
            "Select files to include and set order.\n"
            "Order in this list controls left-to-right placement in each output category sheet."
        ),
        justify="left",
        anchor="w",
    ).pack(fill="x", padx=10, pady=8)

    frame = tk.Frame(win)
    frame.pack(fill="both", expand=True, padx=10, pady=8)

    listbox = tk.Listbox(frame, selectmode="extended")
    listbox.pack(side="left", fill="both", expand=True)

    sb = tk.Scrollbar(frame, orient="vertical", command=listbox.yview)
    sb.pack(side="right", fill="y")
    listbox.configure(yscrollcommand=sb.set)

    display = [f"{clean_fraction_name(f.name)}   [{f.name}]" for f in files]
    for item in display:
        listbox.insert("end", item)

    listbox.selection_set(0, "end")

    result = {"selected": list(files)}

    def move(delta: int):
        sel = list(listbox.curselection())
        if not sel:
            return
        if delta < 0 and sel[0] == 0:
            return
        if delta > 0 and sel[-1] == listbox.size() - 1:
            return

        texts = [listbox.get(i) for i in range(listbox.size())]
        selected_flags = [i in sel for i in range(len(texts))]

        for i in (sel if delta < 0 else reversed(sel)):
            j = i + delta
            texts[i], texts[j] = texts[j], texts[i]
            selected_flags[i], selected_flags[j] = selected_flags[j], selected_flags[i]

        listbox.delete(0, "end")
        for t in texts:
            listbox.insert("end", t)
        for i, is_sel in enumerate(selected_flags):
            if is_sel:
                listbox.selection_set(i)

    def ok():
        selected_idx = listbox.curselection()
        if not selected_idx:
            messagebox.showwarning("No files selected", "Please select at least one input file.", parent=win)
            return
        ordered_text = [listbox.get(i) for i in range(listbox.size())]
        text_to_file = {display[i]: files[i] for i in range(len(files))}
        result["selected"] = [text_to_file[t] for t in ordered_text if listbox.selection_includes(ordered_text.index(t))]
        win.destroy()

    def cancel():
        result["selected"] = []
        win.destroy()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=8)
    tk.Button(btn, text="Move Up", command=lambda: move(-1)).pack(side="left")
    tk.Button(btn, text="Move Down", command=lambda: move(1)).pack(side="left", padx=5)
    tk.Button(btn, text="Select All", command=lambda: listbox.selection_set(0, "end")).pack(side="left", padx=20)
    tk.Button(btn, text="Clear", command=lambda: listbox.selection_clear(0, "end")).pack(side="left")
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.grab_set()
    win.wait_window()
    return result["selected"]


def discover_categories(files: Sequence[Path], logger: LiveLogger) -> List[str]:
    cats = set()
    for p in files:
        try:
            xls = pd.ExcelFile(p)
            for s in xls.sheet_names:
                if s not in IGNORE_SHEETS:
                    cats.add(s)
            logger.write(f"Read sheets from: {p.name}")
        except Exception as exc:
            logger.write(f"WARNING: Could not read {p.name}: {exc}")
    return sorted(cats, key=lambda x: x.lower())


def choose_categories_gui(root: tk.Tk, categories: Sequence[str]) -> List[str]:
    win = tk.Toplevel(root)
    win.title("Select mineral categories to include")
    win.geometry("900x650")

    tk.Label(win, text="Select/deselect category sheets for the summary output.", anchor="w").pack(fill="x", padx=10, pady=8)

    canvas = tk.Canvas(win)
    frame = tk.Frame(canvas)
    sb = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    vars_ = []
    for i, cat in enumerate(categories):
        v = tk.BooleanVar(value=True)
        vars_.append(v)
        tk.Checkbutton(frame, variable=v, text=cat, anchor="w", width=80, justify="left").grid(row=i, column=0, sticky="w")

    result = {"cats": list(categories)}

    def select_all():
        for v in vars_:
            v.set(True)

    def select_none():
        for v in vars_:
            v.set(False)

    def ok():
        chosen = [cat for cat, v in zip(categories, vars_) if v.get()]
        if not chosen:
            messagebox.showwarning("No categories selected", "Please select at least one category.", parent=win)
            return
        result["cats"] = chosen
        win.destroy()

    def cancel():
        result["cats"] = []
        win.destroy()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=8)
    tk.Button(btn, text="Select All", command=select_all).pack(side="left")
    tk.Button(btn, text="Select None", command=select_none).pack(side="left", padx=5)
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.grab_set()
    win.wait_window()
    return result["cats"]


def extract_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    cols_map = {str(c).strip().lower(): c for c in df.columns}
    found = []
    missing = []
    for wanted in REQUIRED_HEADERS:
        k = wanted.lower()
        if k in cols_map:
            found.append(cols_map[k])
        else:
            missing.append(wanted)
    if not found:
        return pd.DataFrame(columns=REQUIRED_HEADERS), missing
    out = df[found].copy()
    out.columns = [h for h in REQUIRED_HEADERS if h.lower() in cols_map]
    return out, missing


def format_block_sheet(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 34)


def write_category_sheet(wb: Workbook, category: str, selected_files: Sequence[Path], logger: LiveLogger):
    ws = wb.create_sheet(title=category[:31])
    row_start = 1 + HEADER_BLANK_ROWS
    current_col = 1

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_font = Font(bold=True)

    for file_path in selected_files:
        file_label = clean_fraction_name(file_path.name)
        ws.cell(row=1, column=current_col, value=file_label)
        ws.cell(row=1, column=current_col).font = Font(bold=True, size=11)

        try:
            if category not in pd.ExcelFile(file_path).sheet_names:
                ws.cell(row=row_start, column=current_col, value=f"Sheet '{category}' not present in this file")
                current_col += len(REQUIRED_HEADERS) + BLOCK_GAP_COLUMNS
                continue

            df_cat = pd.read_excel(file_path, sheet_name=category)
            picked, missing = extract_columns(df_cat)

            for i, h in enumerate(picked.columns, start=0):
                cell = ws.cell(row=row_start, column=current_col + i, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row_vals in enumerate(picked.itertuples(index=False), start=row_start + 1):
                for c_idx, val in enumerate(row_vals, start=0):
                    ws.cell(row=r_idx, column=current_col + c_idx, value=val)

            if missing:
                msg_row = row_start + 1
                ws.cell(row=msg_row, column=current_col + len(picked.columns), value="Missing:")
                ws.cell(row=msg_row, column=current_col + len(picked.columns) + 1, value=", ".join(missing))

            logger.write(f"{category} <- {file_path.name}: {len(picked)} rows")

        except Exception as exc:
            ws.cell(row=row_start, column=current_col, value=f"Error reading file/sheet: {exc}")
            logger.write(f"ERROR in {category} from {file_path.name}: {exc}")

        current_col += len(REQUIRED_HEADERS) + BLOCK_GAP_COLUMNS

    format_block_sheet(ws)


def write_area_sheet(wb: Workbook, selected_files: Sequence[Path], logger: LiveLogger):
    ws = wb.create_sheet(title="Area")
    current_col = 1
    header_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    for file_path in selected_files:
        label = clean_fraction_name(file_path.name)
        ws.cell(row=1, column=current_col, value=label)
        ws.cell(row=1, column=current_col).font = Font(bold=True, size=11)

        try:
            if "Area" not in pd.ExcelFile(file_path).sheet_names:
                ws.cell(row=2, column=current_col, value="Area sheet not present")
                current_col += 12
                continue

            df_area = pd.read_excel(file_path, sheet_name="Area")
            start_row = 1 + HEADER_BLANK_ROWS

            for c, h in enumerate(df_area.columns, start=current_col):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for r, vals in enumerate(df_area.itertuples(index=False), start=start_row + 1):
                for c_off, val in enumerate(vals):
                    ws.cell(row=r, column=current_col + c_off, value=val)

            logger.write(f"Area <- {file_path.name}: {len(df_area)} rows")
            current_col += max(12, len(df_area.columns) + BLOCK_GAP_COLUMNS)

        except Exception as exc:
            ws.cell(row=2, column=current_col, value=f"Error reading Area sheet: {exc}")
            logger.write(f"ERROR Area from {file_path.name}: {exc}")
            current_col += 12

    format_block_sheet(ws)


def main():
    root = tk.Tk()
    root.withdraw()
    logger = LiveLogger(root)

    sample_dir = choose_sample_folder(root)
    if not sample_dir:
        logger.write("No sample folder selected. Exiting.")
        root.destroy()
        return

    logger.write(f"Sample folder: {sample_dir}")
    files = find_excel_files(sample_dir)
    if not files:
        messagebox.showerror("No Excel files", "No Excel files found in selected folder.", parent=root)
        root.destroy()
        return

    logger.write(f"Found {len(files)} Excel files")
    selected_files = choose_file_order_gui(root, files)
    if not selected_files:
        logger.write("No input files selected. Exiting.")
        root.destroy()
        return

    logger.write("Selected file order:")
    for idx, p in enumerate(selected_files, start=1):
        logger.write(f"  {idx}. {p.name}")

    categories = discover_categories(selected_files, logger)
    if not categories:
        messagebox.showerror("No category sheets", "No usable category sheets were detected.", parent=root)
        root.destroy()
        return

    chosen_categories = choose_categories_gui(root, categories)
    if not chosen_categories:
        logger.write("No categories chosen. Exiting.")
        root.destroy()
        return

    out_default = sample_dir / f"{sample_dir.name}_Summary_Each_Sample.xlsx"
    out_path = filedialog.asksaveasfilename(
        parent=root,
        title="Save Summary workbook",
        initialdir=str(sample_dir),
        initialfile=out_default.name,
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
    )
    if not out_path:
        logger.write("No output path selected. Exiting.")
        root.destroy()
        return

    wb = Workbook()
    wb.remove(wb.active)

    for cat in chosen_categories:
        write_category_sheet(wb, cat, selected_files, logger)

    write_area_sheet(wb, selected_files, logger)

    wb.save(out_path)
    logger.write(f"Saved: {out_path}")
    messagebox.showinfo("Done", f"Summary workbook created:\n{out_path}", parent=root)

    try:
        logger.win.destroy()
    except Exception:
        pass
    root.destroy()


if __name__ == "__main__":
    main()
