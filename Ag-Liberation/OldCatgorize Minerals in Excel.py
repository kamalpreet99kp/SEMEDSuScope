import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os

# Mineral classification function
def classify_mineral(row):
    ag = row.get("Ag (Wt%)", 0)
    s = row.get("S (Wt%)", 0)
    hg = row.get("Hg (Wt%)", 0)
    te = row.get("Te (Wt%)", 0)
    se = row.get("Se (Wt%)", 0)
    sb = row.get("Sb (Wt%)", 0)
    ars = row.get("As (Wt%)", 0)
    fe = row.get("Fe (Wt%)", 0)
    cu = row.get("Cu (Wt%)", 0)
    bi = row.get("Bi (Wt%)", 0)
    se = row.get("Se (Wt%)", 0)

    if ag > 10 and s > 7 and hg < 7 and te < 6 and se < 6 and bi < 40 and sb < 8 and ars < 8 and cu < 25 and fe < 26 and se < 6:
        return "Acanthite & Associations"
    elif 5 < ag < 75 and 5 < s < 31 and 6.99 < hg < 53 and te < 6 and sb < 15 and ars < 15 and bi < 40 and cu < 25 and fe < 26 and se < 6:
        return "Imiterite & Associations"
    elif 5 < ag < 75 and s > 7.0 and hg < 12 and fe < 21 and te < 6 and ars > 2 and bi < 40 and cu < 25 and se < 6:
        return "Sulfosalt (Sb & or As ) & Asso"
    elif 5 < ag < 75 and s > 4 and hg < 12 and te < 6 and sb > 6 and ars < 2 and bi < 40 and cu < 25 and fe < 21 and se < 6:
        return "Sulfosalt (Sb) & Asso"
    elif ag > 1.0 and s > 15 and hg < 6 and te < 6 and ars > 10 and cu > 17 and bi < 40 and se < 6:
        return "Ag Associations with Enargite"
    elif ag > 1.4 and hg > 3 and fe < 20 and te < 6 and bi < 40 and cu < 20 and se < 6:
        return "Other Ag-Hg Associations"
    elif ag > 2 and s > 2.5 and hg < 4 and te < 6 and fe < 28 and bi < 40 and cu < 28 and se >= 6:
        return "Aguilarite & Associations"
    elif ag > 5 and te >= 6 and hg < 4 and fe < 28 and bi < 40 and cu < 28:
        return "Hessite & Associations"
    elif ag > 1.4 and (fe > 10 or cu > 10) and bi < 40:
        return "Ag Associations with Sulphides"
    elif ag > 1.4:
        return "All Others"
    else:
        return None

# Required output sheet order
sheet_order = [
    "Area",
    "Summary",
    "Raw Data",
    "Acanthite & Associations",
    "Imiterite & Associations",
    "Sulfosalt (Sb & or As ) & Asso",
    "Sulfosalt (Sb) & Asso",
    "Ag Associations with Enargite",
    "Other Ag-Hg Associations",
    "Aguilarite & Associations",
    "Hessite & Associations",
    "Ag Associations with Sulphides",
    "All Others"
]

# Start file selection dialog
tk.Tk().withdraw()
file_path = filedialog.askopenfilename(title="Select Full Analysis Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])

if not file_path:
    print("No file selected.")
    exit()

# Load workbook and detect the single sheet
df_raw = pd.read_excel(file_path, sheet_name=None)
sheet_name = list(df_raw.keys())[0]
df = df_raw[sheet_name].copy()

# Classify each row
df["Mineral Type"] = df.apply(classify_mineral, axis=1)

# Prepare classified sheets
classified_sheets = {}
area_column = "Area (sq. µm)"
summary_data = []

# Prepare all mineral category sheets
mineral_categories = sheet_order[3:]
for category in mineral_categories:
    df_mineral = df[df["Mineral Type"] == category]
    classified_sheets[category] = df_mineral.drop(columns=["Mineral Type"])
    area_sum = df_mineral[area_column].sum()
    summary_data.append((category, area_sum))

# Compute total area for percentage
total_area = sum(x[1] for x in summary_data)
summary_df = pd.DataFrame(summary_data, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
summary_df["Percentage"] = summary_df["Total Sum of Area (sq. µm)"] / total_area * 100

# Define output filename
output_path = os.path.join(os.path.dirname(file_path), f"Classified_{os.path.basename(file_path)}")

# Write Excel output
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)  # Empty Area sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)
    for category in mineral_categories:
        classified_sheets[category].to_excel(writer, sheet_name=category, index=False)
    # === VERIFICATION CHECK ===
    # Count and area from Raw Data (Ag > 1%)
    df_ag = df[df["Ag (Wt%)"] > 1.4]
    raw_ag_count = len(df_ag)
    raw_ag_area = df_ag[area_column].sum()

    # Count and area from classified sheets
    classified_total_count = sum(len(classified_sheets[cat]) for cat in mineral_categories)
    classified_total_area = sum(classified_sheets[cat][area_column].sum() for cat in mineral_categories)

    # Build integrity check DataFrame
    integrity_data = {
        "Metric": [
            "Rows with Ag > 1.4% in Raw Data",
            "Total rows in classified sheets",
            "Area in Raw Data (Ag > 1.4%)",
            "Area in classified sheets",
            "Area Match",
            "Row Count Match"
        ],
        "Value": [
            raw_ag_count,
            classified_total_count,
            round(raw_ag_area, 4),
            round(classified_total_area, 4),
            "✅ Match" if abs(raw_ag_area - classified_total_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_ag_count == classified_total_count else "❌ Mismatch"
        ]
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)


print(f"\n✅ Classification complete.\nOutput saved to:\n{output_path}")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Color map for each target column
highlight_colors = {
    "Feature": "00BFCF",        # Medium Cyan
    "Area (sq. µm)": "FFEB3B",  # Medium Yellow
    "S (Wt%)": "FFA07A",        # Orange Light (Light Salmon)
    "Fe (Wt%)": "ADD8E6",       # Light Blue
    "Cu (Wt%)": "90EE90",       # Light Green
    "As (Wt%)": "D87093",       # Medium Pink (Pale Violet Red)
    "Ag (Wt%)": "F44336",       # Medium Red
    "Sb (Wt%)": "A9A9A9",       # Medium Grey (Dark Grey)
    "Hg (Wt%)": "9370DB",       # Medium Purple
    "Se (Wt%)": "FFA09A",
    "Te (Wt%)": "D87099",
}

# Border style
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Reopen the saved Excel file
wb = load_workbook(output_path)

# Apply highlighting to all mineral sheets
for sheet in wb.sheetnames:
    if sheet in ["Area", "Integrity Check"]:
        continue
    ws = wb[sheet]
    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Format "Summary" sheet
if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for cell in ws["A"]:
        cell.font = Font(bold=True)

# Save the final formatted workbook
wb.save(output_path)
