import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


def v(row, col):
    x = row.get(col, 0)
    return 0 if pd.isna(x) else x


TRACE = 2.0
MINOR = 6.0
AUAG_CUTOFF = 2.0


def classify_mineral(row):
    ag = v(row, "Ag (Wt%)")
    s = v(row, "S (Wt%)")
    hg = v(row, "Hg (Wt%)")
    te = v(row, "Te (Wt%)")
    se = v(row, "Se (Wt%)")
    sb = v(row, "Sb (Wt%)")
    ars = v(row, "As (Wt%)")
    fe = v(row, "Fe (Wt%)")
    cu = v(row, "Cu (Wt%)")
    bi = v(row, "Bi (Wt%)")
    au = v(row, "Au (Wt%)")
    pb = v(row, "Pb (Wt%)")
    o = v(row, "O (Wt%)")

    # 1) Native Au
    if (30 <= au <= 100 and ag < 1 and te < 2 and bi < 2 and
            pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Native Au"

    # 2) Native Ag
    if (ag >= 85 and o <= 20 and au < MINOR and
            s < TRACE and te < TRACE and se < TRACE and sb < TRACE and ars < TRACE and
            hg < TRACE and bi < TRACE and pb < MINOR and cu < TRACE and fe < TRACE):
        return "Native Ag"

    # 3) Au+Ag (Electrum)
    if (5 <= au <= 98 and 5 <= ag <= 98 and te < 2 and bi < 2 and
            pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Au+Ag (Electrum)"

    # 4) Maldonite
    if (35 <= au <= 70 and 18 <= bi <= 70 and ag < 6 and te < 6 and
            pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Maldonite"

    # 5) Calaverite
    if (35 <= au <= 50 and 45 <= te <= 63 and ag < 6 and bi < 6 and
            pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Calaverite"

    # 6) Sylvanite/Krennerite
    if (49 <= te <= 60 and 17 <= au <= 39 and 3 <= ag <= 15 and
            bi < MINOR and pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and
            sb < TRACE and ars < TRACE and cu < TRACE and fe < TRACE):
        return "Sylvanite/Krennerite"

    # 7) Petzite
    if (30 <= ag <= 48 and 27 <= te <= 36 and 17 <= au <= 29 and
            bi < MINOR and pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and
            sb < TRACE and ars < TRACE and cu < TRACE and fe < TRACE):
        return "Petzite"

    # 8) Kostovite
    if (17 <= au <= 29 and 52 <= te <= 70 and 3.5 <= cu <= 12 and
            ag < MINOR and bi < MINOR and pb < MINOR and s < TRACE and se < TRACE and
            hg < TRACE and sb < TRACE and ars < TRACE and fe < TRACE):
        return "Kostovite"

    # 9) Muthmannite
    if (28 <= au <= 37 and 15 <= ag <= 22 and 40 <= te <= 49 and
            bi < MINOR and pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and
            sb < TRACE and ars < TRACE and cu < TRACE and fe < TRACE):
        return "Muthmannite"

    # 10) Nagyagite
    if (4 <= au <= 10 and 10 <= te <= 18 and 48 <= pb <= 64 and 3 <= sb <= 8 and 5 <= s <= 14 and
            ag < MINOR and bi < MINOR and se < TRACE and hg < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE):
        return "Nagyagite"

    # 11) Hessite
    if (40 <= ag <= 68 and 18 <= te <= 45 and au < 5 and
            s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
            bi < MINOR and pb < MINOR and cu < MINOR and fe < MINOR):
        return "Hessite"

    # 12) Dyscrasite
    if (60 <= ag <= 76 and 20 <= sb <= 30 and s < 5 and
            te < TRACE and se < TRACE and au < TRACE and hg < TRACE and bi < TRACE and
            pb < MINOR and ars < TRACE and cu < TRACE and fe < TRACE):
        return "Dyscrasite"

    # 13) Pyrostilpnite/Stephanite
    if (50 <= ag <= 70 and 12 <= sb <= 25 and 7 <= s <= 20 and
            te < TRACE and se < TRACE and au < TRACE and hg < TRACE and bi < MINOR and
            pb < MINOR and ars < TRACE and cu < TRACE and fe < TRACE):
        return "Pyrostilpnite/Stephanite"

    # 14) Acanthite
    if ag > 10 and s > 7 and hg < 7 and te < 6 and se < 6 and bi < 40 and sb < 8 and ars < 8 and cu < 25 and fe < 26:
        return "Acanthite"

    # 15) Imiterite
    if 5 < ag < 75 and 5 < s < 31 and 6.99 < hg < 53 and te < 6 and sb < 15 and ars < 15 and bi < 40 and cu < 25 and fe < 26 and se < 6:
        return "Imiterite"

    # 16) Sulfosalt (combined)
    if ((5 < ag < 75 and s > 7.0 and hg < 12 and fe < 21 and te < 6 and ars > 2 and bi < 40 and cu < 25 and se < 6) or
            (5 < ag < 75 and s > 4 and hg < 12 and te < 6 and sb > 6 and ars < 2 and bi < 40 and cu < 25 and fe < 21 and se < 6)):
        return "Sulfosalt (Sb/As) & Associations"

    # 17) Aguilarite
    if ag > 2 and s > 2.5 and hg < 4 and te < 6 and fe < 28 and bi < 40 and cu < 28 and se >= 6:
        return "Aguilarite"

    # 18) Bohdanowiczite
    if (se >= 6 and bi >= 15 and te < TRACE and s < TRACE and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= se <= 85 and
            au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE and pb < MINOR):
        return "Bohdanowiczite"

    # 19) Volynskite
    if (te >= 6 and bi >= 15 and se < TRACE and s < TRACE and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= te <= 90 and
            au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
            cu < TRACE and fe < TRACE and pb < MINOR):
        return "Volynskite"

    # 20) Altaite
    if (49 <= pb <= 65 and 29 <= te <= 41 and au < MINOR and ag < MINOR and
            bi < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Altaite"

    # 21) Coloradoite
    if (49 <= hg <= 65 and 29 <= te <= 41 and au < MINOR and ag < MINOR and
            bi < MINOR and pb < MINOR and s < TRACE and se < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Coloradoite"

    # 22) Tellurobismuthite
    if (38 <= te <= 52 and 39 <= bi <= 59 and au < MINOR and ag < MINOR and
            pb < MINOR and s < TRACE and se < TRACE and hg < TRACE and sb < TRACE and
            ars < TRACE and cu < TRACE and fe < TRACE):
        return "Tellurobismuthite"

    # 23) Au Associations with Sulphides
    if au > AUAG_CUTOFF and (s > 5 or fe > 10 or cu > 10):
        return "Au Associations with Sulphides"

    # 24) Ag Associations with Sulphides
    if ag > AUAG_CUTOFF and (s > 5 or fe > 10 or cu > 10):
        return "Ag Associations with Sulphides"

    # 25) All Other (Au)
    if au > AUAG_CUTOFF:
        return "All Other (Au)"

    # 26) All Others (Ag)
    if ag > AUAG_CUTOFF:
        return "All Others (Ag)"

    # 27) All Others (Pb, Bi, Hg)
    if pb > AUAG_CUTOFF or bi > AUAG_CUTOFF or hg > AUAG_CUTOFF:
        return "All Others (Pb, Bi, Hg)"

    return "Unidentified"


tk.Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Select Full Analysis Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    print("No file selected.")
    raise SystemExit

book = pd.read_excel(file_path, sheet_name=None)
raw_sheet_name = list(book.keys())[0]
df = book[raw_sheet_name].copy()

df["Mineral Type"] = df.apply(classify_mineral, axis=1)
area_column = "Area (sq. µm)"

all_categories_in_order = [
    "Native Au", "Native Ag", "Au+Ag (Electrum)", "Maldonite", "Calaverite",
    "Sylvanite/Krennerite", "Petzite", "Kostovite", "Muthmannite", "Nagyagite",
    "Hessite", "Dyscrasite", "Pyrostilpnite/Stephanite", "Acanthite", "Imiterite",
    "Sulfosalt (Sb/As) & Associations", "Aguilarite", "Bohdanowiczite", "Volynskite",
    "Altaite", "Coloradoite", "Tellurobismuthite", "Au Associations with Sulphides",
    "Ag Associations with Sulphides", "All Other (Au)", "All Others (Ag)",
    "All Others (Pb, Bi, Hg)", "Unidentified"
]

classified_sheets = {}
summary_rows = []

for cat in all_categories_in_order:
    df_cat = df[df["Mineral Type"] == cat]
    if len(df_cat) == 0:
        continue
    classified_sheets[cat] = df_cat.drop(columns=["Mineral Type"])
    summary_rows.append((cat, df_cat[area_column].sum()))

summary_df = pd.DataFrame(summary_rows, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
total_area = summary_df["Total Sum of Area (sq. µm)"].sum()
summary_df["Percentage"] = (summary_df["Total Sum of Area (sq. µm)"] / total_area * 100) if total_area else 0

output_path = os.path.join(os.path.dirname(file_path), f"Classified_Tellurides_{os.path.basename(file_path)}")

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)

    for cat in all_categories_in_order:
        if cat in classified_sheets:
            classified_sheets[cat].to_excel(writer, sheet_name=cat, index=False)

    df_focus = df[(df["Ag (Wt%)"] > AUAG_CUTOFF) | (df["Au (Wt%)"] > AUAG_CUTOFF) |
                  (df.get("Pb (Wt%)", 0) > AUAG_CUTOFF) | (df["Bi (Wt%)"] > AUAG_CUTOFF) |
                  (df["Hg (Wt%)"] > AUAG_CUTOFF)]

    raw_count = len(df_focus)
    raw_area = df_focus[area_column].sum()
    classified_rows = df[df["Mineral Type"].notna()]
    classified_count = len(classified_rows)
    classified_area = classified_rows[area_column].sum()

    integrity_data = {
        "Metric": [
            f"Rows with Ag/Au/Pb/Bi/Hg > {AUAG_CUTOFF}% in Raw Data",
            "Total rows in classified sheets (non-empty only)",
            "Area in Raw Data (focused elements)",
            "Area in classified sheets (non-empty only)",
            "Area Match",
            "Row Count Match"
        ],
        "Value": [
            raw_count,
            classified_count,
            round(raw_area, 4),
            round(classified_area, 4),
            "✅ Match" if abs(raw_area - classified_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_count == classified_count else "❌ Mismatch"
        ]
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)

print(f"\n✅ Classification complete.\nOutput saved to:\n{output_path}")

highlight_colors = {
    "Feature": "00BFCF",
    "Area (sq. µm)": "FFEB3B",
    "S (Wt%)": "FFA07A",
    "Fe (Wt%)": "ADD8E6",
    "Cu (Wt%)": "90EE90",
    "As (Wt%)": "D87093",
    "Ag (Wt%)": "F44336",
    "Sb (Wt%)": "A9A9A9",
    "Hg (Wt%)": "9370DB",
    "Se (Wt%)": "FFA09A",
    "Te (Wt%)": "D87099",
    "Au (Wt%)": "FFF2CC",
    "Bi (Wt%)": "E2EFDA",
    "Pb (Wt%)": "F4B183",
    "O (Wt%)": "D9E1F2",
}

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = load_workbook(output_path)

for sheet in wb.sheetnames:
    if sheet in ["Area", "Integrity Check"]:
        continue
    ws = wb[sheet]
    if ws.max_row < 2:
        continue

    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = fill

if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for cell in ws["A"]:
        cell.font = Font(bold=True)

wb.save(output_path)
