import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


# ============================================================
#  UPDATED AG MINERAL CLASSIFICATION - 41-19
#  - Ag-only classification using Ag > 1.4 wt% focus
#  - Clean/free minerals are classified before association groups
#  - Major sections: native/electrum, tellurides, selenides, sulfides,
#    Ag-Hg, sulfosalts, galena associations, and fallback associations
#  - Empty category sheets are not written
#  - Output is written as Classified<original filename>.xlsx
# ============================================================

AG_CUTOFF = 1.4
TRACE_NATIVE = 2.0
TRACE_STRICT = 3.0
TRACE = 4.0
TELLURIDE_GATE = 5.0
SELENIDE_ASSO_GATE = 3.5
SULFIDE_GATE = 4.0
PB_CLEAN_CAP = 15.0
PB_ASSO_CAP = 20.0


def v(row, col):
    """Return a safe numeric value for an optional composition column."""
    x = row.get(col, 0)
    return 0 if pd.isna(x) else x


def excel_safe_sheet_name(name):
    """Excel sheets cannot contain: \\ / ? * [ ] : and max length is 31."""
    cleaned = re.sub(r'[\\/*?:\[\]]', "-", str(name)).strip()
    return cleaned[:31] if cleaned else "Sheet"


def has_required_columns(df, required_columns):
    return [col for col in required_columns if col not in df.columns]


def classify_mineral(row):
    ag = v(row, "Ag (Wt%)")
    s = v(row, "S (Wt%)")
    hg = v(row, "Hg (Wt%)")
    te = v(row, "Te (Wt%)")
    se = v(row, "Se (Wt%)")
    sb = v(row, "Sb (Wt%)")
    ars = v(row, "As (Wt%)")
    fe = v(row, "Fe (Wt%)")
    cu = v(row, "Cu (Wt%)")
    bi = v(row, "Bi (Wt%)")
    au = v(row, "Au (Wt%)")
    pb = v(row, "Pb (Wt%)")
    o = v(row, "O (Wt%)")

    if ag <= AG_CUTOFF and au < 30:
        return None

    # ============================================================
    # 1) NATIVE AG / AU-AG ALLOY - clean categories only
    # ============================================================
    if (au >= 30 and (ag < TRACE_NATIVE or ag > 2) and
            s < TRACE and te < TRACE and se < TRACE and
            ars < TRACE and sb < TRACE and hg < TRACE and bi < TRACE and
            pb < TRACE and cu < TRACE and fe < TRACE):
        return "Au-Ag alloy"

    if (ag >= 82 and o <= 25 and s < TRACE_NATIVE and te < TRACE_NATIVE and
            se < TRACE_NATIVE and ars < TRACE_NATIVE and sb < TRACE_NATIVE and
            hg < TRACE_NATIVE and bi < TRACE_NATIVE and pb < 5 and
            cu < TRACE_NATIVE and fe < TRACE_STRICT and au < TRACE_NATIVE):
        return "Native Ag"

    # ============================================================
    # 2) AG TELLURIDES - clean minerals first, association fallback later
    # ============================================================
    if (30 <= ag <= 48 and 27 <= te <= 36 and 17 <= au <= 29 and
            s < TRACE_NATIVE and se < TRACE_NATIVE and hg < TRACE_NATIVE and
            sb < TRACE_NATIVE and ars < TRACE_NATIVE and bi < 6 and
            pb < 6 and cu < TRACE_NATIVE and fe < TRACE_NATIVE):
        return "Petzite"

    if (40 <= ag <= 68 and 18 <= te <= 45 and au < 5 and
            s < TRACE_NATIVE and se < TRACE_NATIVE and hg < TRACE_NATIVE and
            sb < TRACE_NATIVE and ars < TRACE_NATIVE and bi < 6 and
            pb < 6 and cu < 6 and fe < 6):
        return "Hessite"

    if (te >= 6 and bi >= 15 and se < 5 and s < 5 and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= te <= 90 and
            au < 5 and hg < 5 and sb < 5 and ars < 5 and cu < 5 and
            fe < 5 and pb < PB_CLEAN_CAP):
        return "Volynskite"

    # ============================================================
    # 3) AG SELENIDES / AG-SE-S MINERALS - clean minerals first
    # ============================================================
    if (se >= 6 and bi >= 15 and te < 5 and s < 5 and
            3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= se <= 85 and
            au < 5 and hg < 5 and sb < 5 and ars < 5 and cu < 5 and
            fe < 5 and pb < PB_CLEAN_CAP):
        return "Bohdanowiczite"

    if (55 <= ag <= 85 and 7 <= se <= 20 and 2 <= s <= 10 and
            te < 5 and hg < 4 and ars < 4 and sb < 4 and bi < 6 and
            pb < 6 and cu < 6 and fe < 6 and au < 5):
        return "Aguilarite"

    # ============================================================
    # 4) AG SULFOSALTS - clean minerals before broad sulfide buckets
    # ============================================================
    if (45 <= ag <= 70 and 8 <= s <= 22 and 7 <= ars <= 18 and
            sb < 5 and cu < 6 and hg < 5 and te < TRACE_STRICT and
            se < TRACE_STRICT and bi < 6 and pb < 6 and fe < 6 and au < 5):
        return "Proustite-Xanthoconite"

    if (40 <= ag <= 65 and 8 <= s <= 20 and 8 <= cu <= 23 and
            1 <= ars <= 8 and sb <= 3 and hg < 5 and te < TRACE_STRICT and
            se < TRACE_STRICT and bi < 6 and pb < 6 and fe < 8 and au < 5):
        return "Cupropearceite"

    if (45 <= ag <= 73 and 8 <= sb <= 18 and 7 <= s <= 19 and
            ars < 5 and cu < 6 and hg < 5 and te < TRACE_STRICT and
            se < TRACE_STRICT and bi < 6 and pb < 6 and fe < 6 and au < 5):
        return "Stephanite"

    if (ag >= 30 and sb >= 6 and s < 5 and te < 5 and se < 5 and
            ars < 5 and hg < 5 and bi < 5 and au < 5 and cu < 5 and
            fe < 5 and pb < PB_CLEAN_CAP):
        return "Dyscrasite"

    # ============================================================
    # 5) AG SULFIDES / AG-HG - clean minerals
    # ============================================================
    if (60 <= ag <= 92 and 6 <= s <= 16 and au < 4 and te < TRACE_STRICT and
            se < TRACE_STRICT and hg < TRACE_STRICT and ars < TRACE_STRICT and
            sb < TRACE_STRICT and bi < 6 and pb < 6 and cu < 6 and fe < 6):
        return "Acanthite"

    if (25 <= ag <= 52 and 22 <= hg <= 46 and 6 <= s <= 16 and
            te < TRACE_STRICT and se < TRACE_STRICT and ars < 4 and sb < 4 and
            bi < 6 and pb < 6 and cu < 6 and fe < 6 and au < 5):
        return "Imiterite"

    # ============================================================
    # 6) HIGH-PB GALENA ASSOCIATION - should win over broad asso groups
    # ============================================================
    if ag >= AG_CUTOFF and pb >= 30 and s >= 4 and te < 4 and se < 4:
        return "Ag Inclusion-Association-Rim with Galena"

    # ============================================================
    # 7) ASSOCIATION CATEGORIES
    # ============================================================
    if ag > AG_CUTOFF and te >= TELLURIDE_GATE:
        if au >= AG_CUTOFF:
            return "All Other Au-Te"
        return "All Others Ag-Te"

    if (ag >= 3 and se >= SELENIDE_ASSO_GATE and s > 1 and te < 6 and hg < 4 and
            ars < 8 and sb < 8 and bi < 20 and pb < PB_ASSO_CAP and
            cu < 20 and fe < 20):
        return "Aguilarite-Asso"

    if (ag > 10 and hg > 4 and s > 2 and ars < 5 and sb < 5 and
            se < 5 and te < 6 and bi < 40 and cu < 25 and fe < 26 and
            pb < PB_ASSO_CAP):
        return "Imiterite-Asso"

    if (ag >= 20 and s >= 2.5 and se < 2.5 and hg < 2.5 and
            ars < 2.5 and sb < 2.5 and te < 4 and bi < 40 and
            cu < 25 and fe < 26 and pb < PB_ASSO_CAP):
        return "Acanthite-Asso"

    if (ag > AG_CUTOFF and s >= SULFIDE_GATE and te < 5 and se < 5 and
            (ars >= 2 or sb >= 1) and pb < PB_ASSO_CAP):
        return "Ag-Sulfosalt-Asso"

    if ag > AG_CUTOFF and s >= 2.5 and te < 5 and se < 5 and pb < PB_ASSO_CAP:
        return "Ag Associations with Sulphides"

    if ag > AG_CUTOFF:
        return "All Others Ag"

    return None


# ============================================================
#  FILE SELECT
# ============================================================
tk.Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Select Full Analysis Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    print("No file selected.")
    raise SystemExit

book = pd.read_excel(file_path, sheet_name=None)
raw_sheet_name = list(book.keys())[0]
df = book[raw_sheet_name].copy()

required_columns = ["Ag (Wt%)", "Area (sq. µm)"]
missing_columns = has_required_columns(df, required_columns)
if missing_columns:
    raise ValueError(
        "The selected workbook is missing required column(s): " + ", ".join(missing_columns)
    )

df["Mineral Type"] = df.apply(classify_mineral, axis=1)
area_column = "Area (sq. µm)"

all_categories_in_order = [
    # Native / Au-Ag
    "Native Ag",
    "Au-Ag alloy",

    # Ag tellurides
    "Petzite",
    "Hessite",
    "Volynskite",
    "All Other Au-Te",
    "All Others Ag-Te",

    # Ag selenides / Ag-Se-S
    "Bohdanowiczite",
    "Aguilarite",
    "Aguilarite-Asso",

    # Ag sulfosalts and sulfides
    "Proustite-Xanthoconite",
    "Cupropearceite",
    "Stephanite",
    "Dyscrasite",
    "Acanthite",
    "Imiterite",
    "Ag Inclusion-Association-Rim with Galena",
    "Imiterite-Asso",
    "Acanthite-Asso",
    "Ag-Sulfosalt-Asso",
    "Ag Associations with Sulphides",
    "All Others Ag",
]

classified_sheets = {}
summary_rows = []

for cat in all_categories_in_order:
    df_cat = df[df["Mineral Type"] == cat]
    if len(df_cat) == 0:
        continue
    classified_sheets[cat] = df_cat.drop(columns=["Mineral Type"])
    summary_rows.append((cat, df_cat[area_column].sum()))

summary_df = pd.DataFrame(summary_rows, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
total_area = summary_df["Total Sum of Area (sq. µm)"].sum()
summary_df["Percentage"] = (summary_df["Total Sum of Area (sq. µm)"] / total_area * 100) if total_area else 0

output_path = os.path.join(os.path.dirname(file_path), f"Classified{os.path.basename(file_path)}")


# ============================================================
#  WRITE OUTPUT EXCEL
# ============================================================
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)

    for cat in all_categories_in_order:
        if cat in classified_sheets:
            classified_sheets[cat].to_excel(writer, sheet_name=excel_safe_sheet_name(cat), index=False)

    df_focus = df[(df["Ag (Wt%)"] > AG_CUTOFF) | (df.get("Au (Wt%)", 0) >= 30)]
    raw_count = len(df_focus)
    raw_area = df_focus[area_column].sum()
    classified_total_count = sum(len(classified_sheets[k]) for k in classified_sheets.keys())
    classified_total_area = sum(classified_sheets[k][area_column].sum() for k in classified_sheets.keys())

    integrity_data = {
        "Metric": [
            f"Rows with Ag > {AG_CUTOFF}% or Au >= 30% in Raw Data",
            "Total rows in classified sheets (non-empty only)",
            f"Area in Raw Data (Ag > {AG_CUTOFF}% or Au >= 30%)",
            "Area in classified sheets (non-empty only)",
            "Area Match",
            "Row Count Match",
        ],
        "Value": [
            raw_count,
            classified_total_count,
            round(raw_area, 4),
            round(classified_total_area, 4),
            "✅ Match" if abs(raw_area - classified_total_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_count == classified_total_count else "❌ Mismatch",
        ],
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)

print(f"\n✅ Classification complete.\nOutput saved to:\n{output_path}")


# ============================================================
#  FORMATTING / HIGHLIGHTING
# ============================================================
highlight_colors = {
    "Feature": "00BFCF",
    "Area (sq. µm)": "FFEB3B",
    "S (Wt%)": "FFA07A",
    "Fe (Wt%)": "ADD8E6",
    "Cu (Wt%)": "90EE90",
    "As (Wt%)": "D87093",
    "Ag (Wt%)": "F44336",
    "Sb (Wt%)": "A9A9A9",
    "Hg (Wt%)": "9370DB",
    "Se (Wt%)": "FFA09A",
    "Te (Wt%)": "D87099",
    "Au (Wt%)": "FFF2CC",
    "Bi (Wt%)": "E2EFDA",
    "Pb (Wt%)": "F4B183",
    "O (Wt%)": "D9E1F2",
}

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = load_workbook(output_path)

for sheet in wb.sheetnames:
    if sheet in ["Area", "Integrity Check"]:
        continue
    ws = wb[sheet]
    if ws.max_row < 2:
        continue

    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = fill

if "Summary" in wb.sheetnames:
    ws = wb["Summary"]
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for cell in ws["A"]:
        cell.font = Font(bold=True)

wb.save(output_path)
