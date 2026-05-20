from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

IMG_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}
TARGET_IMAGE_HEADER = "Image"
TARGET_MICROX_HEADER = "Micro X"
LINK_HEADER = "Open Image"
NEW_HEADERS = ["Area Orig", "Corr %", "New Brea", "New Len", "Asso"]
MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 28


class LiveLogger:
    def __init__(self, root: tk.Tk):
        self.win = tk.Toplevel(root)
        self.win.title("Image Links + CorrFactor - Live Progress")
        self.win.geometry("980x520")
        self.text = tk.Text(self.win, wrap="word")
        self.text.pack(fill="both", expand=True)
        self.write("=== Live Progress ===")

    def write(self, msg: str):
        self.text.insert("end", msg + "\n")
        self.text.see("end")
        self.win.update_idletasks()
        self.win.update()


def canonical_key(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("&", "_").replace("-", "_")
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def canonical_variants(s: str) -> List[str]:
    base = canonical_key(s)
    if not base:
        return []
    return list({base, base.replace("_", ""), base.replace("_and_", "_"), base.replace("and", "")})


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def clean_key(value: str) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\.[a-z0-9]{2,5}$", "", s)
    s = s.replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def choose_workbook(root: tk.Tk) -> Path:
    p = filedialog.askopenfilename(parent=root, title="Select workbook", filetypes=[("Excel", "*.xlsx *.xlsm")])
    if not p:
        raise RuntimeError("No workbook selected")
    return Path(p)


def resolve_folder_for_sheet(sheet_name: str, folder_map: Dict[str, Path]) -> Optional[Path]:
    k = canonical_key(sheet_name)
    if k in folder_map:
        return folder_map[k]
    for v in canonical_variants(sheet_name):
        if v in folder_map:
            return folder_map[v]
    for fk, fp in folder_map.items():
        if k and (k in fk or fk in k):
            return fp
    return None


def build_category_folder_map(root_dir: Path) -> Dict[str, Path]:
    folder_map = {}
    for p in root_dir.iterdir():
        if p.is_dir() and (p / "cropped").is_dir():
            folder_map[canonical_key(p.name)] = p
            for v in canonical_variants(p.name):
                folder_map.setdefault(v, p)
    return folder_map


def build_cropped_index(cropped_dir: Path):
    exact = {}
    by_clean = defaultdict(list)
    for p in cropped_dir.iterdir():
        if p.is_file() and p.suffix.lower() in IMG_EXTS:
            exact[normalize_text(p.stem)] = p
            by_clean[clean_key(p.stem)].append(p)
    return exact, by_clean


def find_best_image(micro_x_value: str, exact, by_clean):
    if micro_x_value is None:
        return None
    raw = str(micro_x_value).strip()
    if not raw:
        return None
    k1 = normalize_text(Path(raw).stem)
    if k1 in exact:
        return exact[k1]
    k2 = clean_key(raw)
    if k2 in by_clean and by_clean[k2]:
        return sorted(by_clean[k2], key=lambda p: len(p.name))[0]
    for ck, vals in by_clean.items():
        if ck.startswith(k2) or k2.startswith(ck) or k2 in ck:
            return sorted(vals, key=lambda p: len(p.name))[0]
    return None


def get_header_positions(ws) -> Dict[str, List[int]]:
    positions: Dict[str, List[int]] = defaultdict(list)
    for c in ws[1]:
        if c.value is None:
            continue
        positions[str(c.value).strip()].append(c.column)
    return dict(positions)


def choose_pairs_gui(root: tk.Tk, pairs: List[Tuple[str, Path]], logger: LiveLogger) -> List[Tuple[str, Path]]:
    logger.write(f"Sheets with matched folders: {len(pairs)}")
    if not pairs:
        return []

    win = tk.Tk()
    win.title("Select categories to process")
    win.geometry("920x650")

    tk.Label(win, text="Select categories (sheet + folder) to process", anchor="w").pack(fill="x", padx=10, pady=8)

    canvas = tk.Canvas(win)
    frame = tk.Frame(canvas)
    sb = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=sb.set)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    tk.Label(frame, text="Use", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=0, sticky="w")
    tk.Label(frame, text="Sheet", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=1, sticky="w")
    tk.Label(frame, text="Folder", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=2, sticky="w")

    vars_ = []
    for i, (sheet, folder) in enumerate(pairs, start=1):
        v = tk.BooleanVar(value=True)
        vars_.append(v)
        tk.Checkbutton(frame, variable=v).grid(row=i, column=0, sticky="w")
        tk.Label(frame, text=sheet).grid(row=i, column=1, sticky="w", padx=8)
        tk.Label(frame, text=str(folder), wraplength=520, justify="left").grid(row=i, column=2, sticky="w", padx=8)

    result = {"pairs": pairs}

    def select_all():
        for v in vars_:
            v.set(True)

    def select_none():
        for v in vars_:
            v.set(False)

    def ok():
        result["pairs"] = [p for p, v in zip(pairs, vars_) if v.get()]
        win.quit()

    def cancel():
        result["pairs"] = []
        win.quit()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=8)
    tk.Button(btn, text="Select All", command=select_all).pack(side="left")
    tk.Button(btn, text="Select None", command=select_none).pack(side="left", padx=5)
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.protocol("WM_DELETE_WINDOW", cancel)
    win.mainloop()
    try:
        win.destroy()
    except Exception:
        pass
    return result["pairs"]


def shrink_columns_keep_visible(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        if bool(getattr(dim, "hidden", False)):
            continue
        header = ws.cell(row=1, column=col_idx).value
        header_text = str(header).strip() if header is not None else ""
        if header_text == TARGET_IMAGE_HEADER:
            continue
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        first_word = (re.split(r"\s+", header_text)[0] if header_text else "")
        dim.width = max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, len(first_word) + 1))


def freeze_top_header_row(ws):
    # Freeze only the first row (header), no frozen columns.
    ws.freeze_panes = "A2"


def ensure_new_columns_and_formulas(ws, area_orig_col: int, corr_col: int, new_brea_col: int, new_len_col: int, asso_col: int,
                                    breadth_col: int, length_col: int):
    for r in range(2, ws.max_row + 1):
        # Corr % is user editable; leave blank unless already set.
        corr_ref = f"{get_column_letter(corr_col)}{r}"
        area_ref = f"{get_column_letter(area_orig_col)}{r}"
        br_ref = f"{get_column_letter(breadth_col)}{r}"
        ln_ref = f"{get_column_letter(length_col)}{r}"

        ws.cell(row=r, column=new_brea_col).value = f'=IF(OR({corr_ref}="",{corr_ref}=0),{br_ref},SQRT({area_ref}*{corr_ref}/100))'
        ws.cell(row=r, column=new_len_col).value = f'=IF(OR({corr_ref}="",{corr_ref}=0),{ln_ref},SQRT({area_ref}*{corr_ref}/100))'
        # Asso left blank/editable


def process_sheet(ws, cropped_dir: Path):
    headers = get_header_positions(ws)
    if TARGET_IMAGE_HEADER not in headers or TARGET_MICROX_HEADER not in headers:
        return 0, 0, "Missing Image/Micro X headers"

    # Preserve hidden states by header names (more stable than index when inserting columns)
    hidden_by_header = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        h = str(cell.value).strip()
        hidden_by_header.setdefault(h, False)
        if ws.column_dimensions[get_column_letter(cell.column)].hidden:
            hidden_by_header[h] = True

    image_col = headers[TARGET_IMAGE_HEADER][0]
    micro_col = headers[TARGET_MICROX_HEADER][0]

    if LINK_HEADER in headers:
        link_col = headers[LINK_HEADER][0]
    else:
        insert_at = image_col + 1
        ws.insert_cols(insert_at, amount=1 + len(NEW_HEADERS))
        ws.cell(row=1, column=insert_at).value = LINK_HEADER
        for i, h in enumerate(NEW_HEADERS, start=1):
            ws.cell(row=1, column=insert_at + i).value = h
        link_col = insert_at

    # refresh headers after possible insert
    headers = get_header_positions(ws)

    area_cols = headers.get("Area", [])
    if len(area_cols) < 2:
        return 0, 0, "Could not find second Area column"
    area_second_col = area_cols[1]

    breadth_col = headers.get("Breadth (um)", [None])[0]
    length_col = headers.get("Length", [None])[0]
    if breadth_col is None or length_col is None:
        return 0, 0, "Missing Breadth (um) or Length column"

    area_orig_col = headers["Area Orig"][0]
    corr_col = headers["Corr %"][0]
    new_brea_col = headers["New Brea"][0]
    new_len_col = headers["New Len"][0]
    asso_col = headers["Asso"][0]

    for h in [LINK_HEADER] + NEW_HEADERS:
        c = headers[h][0]
        ws.cell(row=1, column=c).font = Font(bold=True)

    exact_index, clean_index = build_cropped_index(cropped_dir)
    linked, missing = 0, 0
    for r in range(2, ws.max_row + 1):
        mx = ws.cell(row=r, column=micro_col).value
        img_path = find_best_image(mx, exact_index, clean_index)
        c = ws.cell(row=r, column=link_col)
        if img_path is None:
            c.value = "Missing"
            c.hyperlink = None
            missing += 1
        else:
            c.value = "Open Image"
            c.hyperlink = img_path.resolve().as_uri()
            c.font = Font(color="0563C1", underline="single")
            c.alignment = Alignment(horizontal="center", vertical="center")
            linked += 1

        ws.cell(row=r, column=area_orig_col).value = ws.cell(row=r, column=area_second_col).value

    ensure_new_columns_and_formulas(ws, area_orig_col, corr_col, new_brea_col, new_len_col, asso_col, breadth_col, length_col)

    freeze_top_header_row(ws)
    shrink_columns_keep_visible(ws)

    # Re-apply hidden by header label
    headers_now = get_header_positions(ws)
    for h, hidden in hidden_by_header.items():
        if not hidden:
            continue
        for col_idx in headers_now.get(h, []):
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    return linked, missing, "OK"


def main():
    root = tk.Tk()
    root.withdraw()
    try:
        logger = LiveLogger(root)
        wb_path = choose_workbook(root)
        logger.write(f"Workbook selected: {wb_path}")

        root_dir = wb_path.parent
        logger.write(f"Auto-using workbook directory as category root: {root_dir}")

        wb = load_workbook(wb_path)
        folder_map = build_category_folder_map(root_dir)
        logger.write(f"Detected category folders with cropped: {len(folder_map)}")

        pairs = []
        for sheet in wb.sheetnames:
            f = resolve_folder_for_sheet(sheet, folder_map)
            if f is not None:
                pairs.append((sheet, f))

        selected = choose_pairs_gui(root, pairs, logger)
        if not selected:
            logger.write("No categories selected. Exiting.")
            messagebox.showwarning("No selection", "No categories selected.", parent=root)
            return

        total_linked = total_missing = processed = 0
        skipped = []
        for sheet, folder in selected:
            logger.write(f"Processing: {sheet}")
            linked, missing, status = process_sheet(wb[sheet], folder / "cropped")
            if status != "OK":
                skipped.append(f"{sheet}: {status}")
                logger.write(f"  Skipped: {status}")
                continue
            processed += 1
            total_linked += linked
            total_missing += missing
            logger.write(f"  Done links={linked}, missing={missing}")

        out = wb_path.with_name(f"{wb_path.stem}_with_quick_links{wb_path.suffix}")
        logger.write(f"Saving: {out}")
        wb.save(out)
        logger.write("Save complete.")

        formula_info = (
            "New Brea/New Len formula:\n"
            "IF(Corr% blank or 0, keep original Breadth/Length, else SQRT(Area Orig * Corr% / 100))"
        )
        summary = (
            f"Saved: {out}\nProcessed sheets: {processed}\nLinks added: {total_linked}\n"
            f"Missing links: {total_missing}\n\n{formula_info}"
        )
        if skipped:
            summary += "\n\nSkipped:\n" + "\n".join(skipped[:20])

        messagebox.showinfo("Done", summary, parent=root)
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
