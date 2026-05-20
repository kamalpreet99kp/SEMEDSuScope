from __future__ import annotations

"""
Add per-row quick-open hyperlinks to an existing Ag liberation workbook.

Workflow:
1) Select one workbook (.xlsx/.xlsm) that already contains category sheets.
2) Select the parent folder that contains category folders with `cropped` subfolders.
3) Select which sheets to process.
4) For each selected sheet having both `Image` and `Micro X`, add/update `Open Image`
   hyperlinks matched ONLY against that sheet's matching `cropped` folder.
5) Save as a new file with suffix `_with_quick_links`.
"""

import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

IMG_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}
TARGET_IMAGE_HEADER = "Image"
TARGET_MICROX_HEADER = "Micro X"
LINK_HEADER = "Open Image"

MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 28
SAMPLE_ROWS_FOR_WIDTH = 500


def canonical_key(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("&", "_")
    s = s.replace("-", "_")
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def clean_key(value: str) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\.[a-z0-9]{2,5}$", "", s)
    s = s.replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def choose_workbook(root: tk.Tk) -> Path:
    selected = filedialog.askopenfilename(
        parent=root,
        title="Select the Ag workbook to update",
        filetypes=[("Excel Workbook", "*.xlsx *.xlsm")],
    )
    if not selected:
        raise RuntimeError("No workbook selected")
    return Path(selected)


def choose_root_folder(root: tk.Tk) -> Path:
    selected = filedialog.askdirectory(
        parent=root,
        title="Select parent folder containing category folders with 'cropped'",
    )
    if not selected:
        raise RuntimeError("No image root folder selected")
    return Path(selected)


def choose_sheets_listbox(root: tk.Tk, sheet_names: List[str]) -> Set[str]:
    win = tk.Toplevel(root)
    win.title("Choose sheets to process")
    win.geometry("520x620")

    tk.Label(
        win,
        text="Select sheet names to process (Ctrl/Shift for multi-select).",
        anchor="w",
        justify="left",
    ).pack(fill="x", padx=10, pady=(10, 5))

    frame = tk.Frame(win)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")

    listbox = tk.Listbox(frame, selectmode="extended", yscrollcommand=scrollbar.set)
    for s in sheet_names:
        listbox.insert("end", s)
    listbox.selection_set(0, "end")
    listbox.pack(side="left", fill="both", expand=True)
    scrollbar.config(command=listbox.yview)

    result = {"value": set(sheet_names)}

    def select_all():
        listbox.selection_set(0, "end")

    def select_none():
        listbox.selection_clear(0, "end")

    def ok():
        sel = listbox.curselection()
        result["value"] = {sheet_names[i] for i in sel}
        win.destroy()

    def cancel():
        result["value"] = set()
        win.destroy()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=(0, 10))
    tk.Button(btn, text="Select All", command=select_all).pack(side="left")
    tk.Button(btn, text="Select None", command=select_none).pack(side="left", padx=5)
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.transient(root)
    win.grab_set()
    root.wait_window(win)
    return result["value"]


def choose_sheet_folder_pairs(root: tk.Tk, sheet_names: List[str], folder_map: Dict[str, Path]) -> List[Tuple[str, Path]]:
    """Show checkbox GUI of sheet<->folder matches and return selected pairs."""
    pairs = []
    for sheet in sheet_names:
        folder = folder_map.get(canonical_key(sheet))
        if folder is not None:
            pairs.append((sheet, folder))

    if not pairs:
        return []

    win = tk.Toplevel(root)
    win.title("Select categories to process")
    win.geometry("860x640")

    tk.Label(
        win,
        text=(
            "Select categories to process. Each row is a workbook sheet matched to a folder.\n"
            "Only checked rows will be processed."
        ),
        anchor="w",
        justify="left",
    ).pack(fill="x", padx=10, pady=(10, 6))

    container = tk.Frame(win)
    container.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = tk.Canvas(container)
    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas)

    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    tk.Label(inner, text="Use", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=0, sticky="w", padx=(4, 12), pady=(2, 4))
    tk.Label(inner, text="Sheet", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=1, sticky="w", padx=(0, 16), pady=(2, 4))
    tk.Label(inner, text="Matched Folder", font=("TkDefaultFont", 9, "bold")).grid(row=0, column=2, sticky="w", pady=(2, 4))

    vars_by_idx = []
    for idx, (sheet, folder) in enumerate(pairs, start=1):
        var = tk.BooleanVar(value=True)
        vars_by_idx.append(var)
        tk.Checkbutton(inner, variable=var).grid(row=idx, column=0, sticky="w", padx=(4, 12))
        tk.Label(inner, text=sheet, anchor="w").grid(row=idx, column=1, sticky="w", padx=(0, 16))
        tk.Label(inner, text=str(folder), anchor="w", justify="left", wraplength=480).grid(row=idx, column=2, sticky="w")

    result = {"pairs": pairs}

    def select_all():
        for v in vars_by_idx:
            v.set(True)

    def select_none():
        for v in vars_by_idx:
            v.set(False)

    def ok():
        chosen = [pair for pair, v in zip(pairs, vars_by_idx) if v.get()]
        result["pairs"] = chosen
        win.destroy()

    def cancel():
        result["pairs"] = []
        win.destroy()

    btn = tk.Frame(win)
    btn.pack(fill="x", padx=10, pady=(0, 10))
    tk.Button(btn, text="Select All", command=select_all).pack(side="left")
    tk.Button(btn, text="Select None", command=select_none).pack(side="left", padx=5)
    tk.Button(btn, text="OK", command=ok).pack(side="right")
    tk.Button(btn, text="Cancel", command=cancel).pack(side="right", padx=5)

    win.transient(root)
    win.grab_set()
    root.wait_window(win)
    return result["pairs"]

def build_category_folder_map(root_dir: Path) -> Dict[str, Path]:
    folder_map = {}
    for p in root_dir.iterdir():
        if not p.is_dir():
            continue
        cropped = p / "cropped"
        if cropped.is_dir():
            folder_map[canonical_key(p.name)] = p
    return folder_map


def build_cropped_index(cropped_dir: Path):
    exact = {}
    by_clean = defaultdict(list)
    for p in cropped_dir.iterdir():
        if p.is_file() and p.suffix.lower() in IMG_EXTS:
            exact[normalize_text(p.stem)] = p
            by_clean[clean_key(p.stem)].append(p)
    return exact, by_clean


def find_best_image(micro_x_value: str, exact, by_clean):
    if micro_x_value is None:
        return None
    raw = str(micro_x_value).strip()
    if not raw:
        return None

    key_exact = normalize_text(Path(raw).stem)
    if key_exact in exact:
        return exact[key_exact]

    key_clean = clean_key(raw)
    if key_clean in by_clean and by_clean[key_clean]:
        return sorted(by_clean[key_clean], key=lambda p: len(p.name))[0]

    for ck, vals in by_clean.items():
        if ck.startswith(key_clean) or key_clean.startswith(ck) or key_clean in ck:
            return sorted(vals, key=lambda p: len(p.name))[0]
    return None


def get_header_map(ws):
    headers = {}
    for c in ws[1]:
        if c.value is not None:
            headers[str(c.value).strip()] = c.column
    return headers


def shrink_columns_keep_visible(ws):
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        header_text = str(header_val).strip() if header_val is not None else ""

        # Wrap header to preserve readability while keeping narrow width.
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header_words = [w for w in re.split(r"\s+", header_text) if w]
        header_word_width = max((len(w) for w in header_words), default=len(header_text))

        data_max = 0
        max_r = min(ws.max_row, SAMPLE_ROWS_FOR_WIDTH)
        for r in range(2, max_r + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            txt = str(v)
            data_max = max(data_max, len(txt))

        desired = max(header_word_width, min(data_max, MAX_COL_WIDTH))
        desired = max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, desired + 1))

        ws.column_dimensions[get_column_letter(col_idx)].width = desired


def add_links_to_sheet(ws, cropped_dir: Path) -> Tuple[int, int, str]:
    headers = get_header_map(ws)
    if TARGET_IMAGE_HEADER not in headers or TARGET_MICROX_HEADER not in headers:
        return 0, 0, "Sheet missing required headers"

    image_col = headers[TARGET_IMAGE_HEADER]
    microx_col = headers[TARGET_MICROX_HEADER]

    if LINK_HEADER in headers:
        link_col = headers[LINK_HEADER]
    else:
        link_col = image_col + 1
        ws.insert_cols(link_col)
        ws.cell(row=1, column=link_col).value = LINK_HEADER

    ws.cell(row=1, column=link_col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(link_col)].width = max(12, len("Open Image") + 2)

    exact_index, clean_index = build_cropped_index(cropped_dir)

    linked = 0
    missing = 0
    for r in range(2, ws.max_row + 1):
        micro_x = ws.cell(row=r, column=microx_col).value
        cell = ws.cell(row=r, column=link_col)

        img_path = find_best_image(micro_x, exact_index, clean_index)
        if img_path is None:
            cell.value = "Missing"
            cell.hyperlink = None
            missing += 1
            continue

        cell.value = "Open Image"
        cell.hyperlink = img_path.resolve().as_uri()
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        linked += 1

    shrink_columns_keep_visible(ws)
    return linked, missing, "OK"


def main():
    root = tk.Tk()
    root.withdraw()
    try:
        wb_path = choose_workbook(root)
        root_dir = choose_root_folder(root)

        wb = load_workbook(wb_path)
        folder_map = build_category_folder_map(root_dir)

        selected_pairs = choose_sheet_folder_pairs(root, wb.sheetnames, folder_map)
        if not selected_pairs:
            messagebox.showwarning("No categories selected", "No sheet/folder categories were selected for processing.", parent=root)
            return


        total_linked = 0
        total_missing = 0
        processed = 0
        skipped: List[str] = []

        for sheet_name, cat_folder in selected_pairs:
            ws = wb[sheet_name]
            cropped_dir = cat_folder / "cropped"
            if not cropped_dir.is_dir():
                skipped.append(f"{sheet_name}: cropped missing")
                continue

            linked, missing, status = add_links_to_sheet(ws, cropped_dir)
            if status != "OK":
                skipped.append(f"{sheet_name}: {status}")
                continue

            processed += 1
            total_linked += linked
            total_missing += missing

        out = wb_path.with_name(f"{wb_path.stem}_with_quick_links{wb_path.suffix}")
        wb.save(out)

        summary = (
            f"Saved: {out}\n"
            f"Categories selected: {len(selected_pairs)}\n"
            f"Sheets processed: {processed}\n"
            f"Links added: {total_linked}\n"
            f"Missing links: {total_missing}"
        )
        if skipped:
            preview = "\n".join(skipped[:12])
            more = "" if len(skipped) <= 12 else f"\n... and {len(skipped)-12} more"
            summary += f"\n\nSkipped:\n{preview}{more}"

        messagebox.showinfo("Done", summary, parent=root)
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
