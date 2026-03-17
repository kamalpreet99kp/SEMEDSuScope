import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


# ============================================================
#  AG MINERAL CLASSIFICATION (YOUR EXISTING RULES + NEW ONES)
#  - Keeps your existing categories/rules unchanged
#  - Adds: Petzite, Bohdanowiczite, Volynskite, Au-Ag, Native Ag, Dyscrasite
#  - Tight "association caps" (mostly <5–6; occasionally <10–12)
#  - Empty category sheets are NOT written
#  - Summary includes only non-empty categories
#  - Integrity Check uses Ag > 1.4 (same as your prior script)
# ============================================================

# ---- small helper: safe numeric ----
def v(row, col):
    x = row.get(col, 0)
    return 0 if pd.isna(x) else x

# ---- Tunable caps (your request: mostly <5–6, sometimes <8–12) ----
TRACE = 5.0     # typical "shouldn't be there" cap
MINOR = 12.0    # occasional "association allowed" cap


def classify_mineral(row):
    ag = v(row, "Ag (Wt%)")
    s  = v(row, "S (Wt%)")
    hg = v(row, "Hg (Wt%)")
    te = v(row, "Te (Wt%)")
    se = v(row, "Se (Wt%)")
    sb = v(row, "Sb (Wt%)")
    ars = v(row, "As (Wt%)")
    fe = v(row, "Fe (Wt%)")
    cu = v(row, "Cu (Wt%)")
    bi = v(row, "Bi (Wt%)")
    au = v(row, "Au (Wt%)")
    o  = v(row, "O (Wt%)")

    # ============================================================
    # NEW / REFINED CATEGORIES WITH TIGHT CAPS
    # (placed early to avoid being swallowed by broad buckets)
    # ============================================================

    # Bohdanowiczite (AgBiSe2) : Se + Bi + Ag, no Te/S, low others
    if (se >= 6 and bi >= 15 and te < TRACE and s < TRACE and
        3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= se <= 85 and
        au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
        cu < TRACE and fe < TRACE):
        return "Bohdanowiczite"

    # Volynskite (AgBiTe2) : Te + Bi + Ag, no Se/S, low others
    if (te >= 6 and bi >= 15 and se < TRACE and s < TRACE and
        3 <= ag <= 60 and 15 <= bi <= 80 and 8 <= te <= 90 and
        au < TRACE and hg < TRACE and sb < TRACE and ars < TRACE and
        cu < TRACE and fe < TRACE):
        return "Volynskite"

    # Petzite (Ag3AuTe2) : Te + Au + Ag, low S/Se/Bi
    # allow Bi slightly (MINOR) because tellurides sometimes sit near Bi phases,
    # but keep it controlled.
    if (te >= 6 and au >= 5 and s < TRACE and se < TRACE and
        10 <= ag <= 75 and 5 <= au <= 85 and 8 <= te <= 90 and
        bi < MINOR and hg < TRACE and sb < TRACE and ars < TRACE and
        cu < TRACE and fe < TRACE):
        return "Petzite"

    # Hessite & Associations (Ag2Te) : Ag + Te, keep others low (your request)
    if (ag > 5 and te >= 6 and
        s < TRACE and se < TRACE and au < TRACE and
        hg < TRACE and sb < TRACE and ars < TRACE and
        cu < MINOR and fe < MINOR and bi < MINOR):
        return "Hessite & Associations"

    # Dyscrasite (Ag3Sb) : Ag + Sb alloy, no S/Te/Se, low others
    if (ag >= 30 and sb >= 6 and
        s < TRACE and te < TRACE and se < TRACE and
        ars < TRACE and hg < TRACE and bi < TRACE and
        au < TRACE and cu < TRACE and fe < TRACE):
        return "Dyscrasite"

    # Au-Ag alloy (Electrum-like) : Au + Ag, exclude Te/S/Se strongly
    if (au >= 20 and ag >= 5 and
        te < TRACE and s < TRACE and se < TRACE and
        sb < TRACE and ars < TRACE and hg < TRACE and bi < TRACE and
        cu < TRACE and fe < TRACE):
        return "Au-Ag"

    # Native Ag : very high Ag; allow O up to ~20; everything else low
    if (ag >= 85 and o <= 20 and
        s < TRACE and te < TRACE and se < TRACE and
        sb < TRACE and ars < TRACE and hg < TRACE and bi < TRACE and
        au < MINOR and cu < TRACE and fe < TRACE):
        return "Native Ag"

    # ============================================================
    # YOUR EXISTING CATEGORIES (UNCHANGED)
    # ============================================================

    # Acanthite & Associations
    if ag > 10 and s > 7 and hg < 7 and te < 6 and se < 6 and bi < 40 and sb < 8 and ars < 8 and cu < 25 and fe < 26 and se < 6:
        return "Acanthite & Associations"

    # Imiterite & Associations
    elif 5 < ag < 75 and 5 < s < 31 and 6.99 < hg < 53 and te < 6 and sb < 15 and ars < 15 and bi < 40 and cu < 25 and fe < 26 and se < 6:
        return "Imiterite & Associations"

    # Sulfosalt (Sb & or As ) & Asso
    elif 5 < ag < 75 and s > 7.0 and hg < 12 and fe < 21 and te < 6 and ars > 2 and bi < 40 and cu < 25 and se < 6:
        return "Sulfosalt (Sb & or As ) & Asso"

    # Sulfosalt (Sb) & Asso
    elif 5 < ag < 75 and s > 4 and hg < 12 and te < 6 and sb > 6 and ars < 2 and bi < 40 and cu < 25 and fe < 21 and se < 6:
        return "Sulfosalt (Sb) & Asso"

    # Ag Associations with Enargite
    elif ag > 1.0 and s > 15 and hg < 6 and te < 6 and ars > 10 and cu > 17 and bi < 40 and se < 6:
        return "Ag Associations with Enargite"

    # Other Ag-Hg Associations
    elif ag > 1.4 and hg > 3 and fe < 20 and te < 6 and bi < 40 and cu < 20 and se < 6:
        return "Other Ag-Hg Associations"

    # Aguilarite & Associations
    elif ag > 2 and s > 2.5 and hg < 4 and te < 6 and fe < 28 and bi < 40 and cu < 28 and se >= 6:
        return "Aguilarite & Associations"

    # Ag Associations with Sulphides  (allows higher Fe/Cu/S by design)
    elif ag > 1.4 and (fe > 10 or cu > 10) and bi < 40:
        return "Ag Associations with Sulphides"

    # All Others
    elif ag > 1.4:
        return "All Others"

    else:
        return None


# ============================================================
#  FILE SELECT
# ============================================================
tk.Tk().withdraw()
file_path = filedialog.askopenfilename(
    title="Select Full Analysis Excel File",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    print("No file selected.")
    raise SystemExit

# Load workbook (single sheet, name varies)
df_raw = pd.read_excel(file_path, sheet_name=None)
raw_sheet_name = list(df_raw.keys())[0]
df = df_raw[raw_sheet_name].copy()

# Classify
df["Mineral Type"] = df.apply(classify_mineral, axis=1)

area_column = "Area (sq. µm)"


# ============================================================
#  CATEGORY ORDER (includes all; empty sheets will be skipped)
# ============================================================
all_categories_in_order = [
    # New refined categories first
    "Bohdanowiczite",
    "Volynskite",
    "Petzite",
    "Hessite & Associations",
    "Dyscrasite",
    "Au-Ag",
    "Native Ag",
    # Existing categories (your script)
    "Acanthite & Associations",
    "Imiterite & Associations",
    "Sulfosalt (Sb & or As ) & Asso",
    "Sulfosalt (Sb) & Asso",
    "Ag Associations with Enargite",
    "Other Ag-Hg Associations",
    "Aguilarite & Associations",
    "Ag Associations with Sulphides",
    "All Others",
]

classified_sheets = {}
summary_rows = []

for cat in all_categories_in_order:
    df_cat = df[df["Mineral Type"] == cat]
    if len(df_cat) == 0:
        continue
    classified_sheets[cat] = df_cat.drop(columns=["Mineral Type"])
    summary_rows.append((cat, df_cat[area_column].sum()))

summary_df = pd.DataFrame(summary_rows, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
total_area = summary_df["Total Sum of Area (sq. µm)"].sum()
summary_df["Percentage"] = (summary_df["Total Sum of Area (sq. µm)"] / total_area * 100) if total_area else 0

output_path = os.path.join(os.path.dirname(file_path), f"Classified_{os.path.basename(file_path)}")


# ============================================================
#  WRITE OUTPUT EXCEL (skip empty mineral sheets)
#  Order: Area, Summary, Raw Data, (non-empty mineral sheets), Integrity Check (last)
# ============================================================
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)  # Empty

    summary_df.to_excel(writer, sheet_name="Summary", index=False)

    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)

    for cat in all_categories_in_order:
        if cat in classified_sheets:
            classified_sheets[cat].to_excel(writer, sheet_name=cat, index=False)

    # === Integrity Check (Ag > 1.4) ===
    df_ag = df[df["Ag (Wt%)"] > 1.4]
    raw_ag_count = len(df_ag)
    raw_ag_area = df_ag[area_column].sum()

    classified_total_count = sum(len(classified_sheets[k]) for k in classified_sheets.keys())
    classified_total_area = sum(classified_sheets[k][area_column].sum() for k in classified_sheets.keys())

    integrity_data = {
        "Metric": [
            "Rows with Ag > 1.4% in Raw Data",
            "Total rows in classified sheets (non-empty only)",
            "Area in Raw Data (Ag > 1.4%)",
            "Area in classified sheets (non-empty only)",
            "Area Match",
            "Row Count Match"
        ],
        "Value": [
            raw_ag_count,
            classified_total_count,
            round(raw_ag_area, 4),
            round(classified_total_area, 4),
            "✅ Match" if abs(raw_ag_area - classified_total_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_ag_count == classified_total_count else "❌ Mismatch"
        ]
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)

print(f"\n✅ Classification complete.\nOutput saved to:\n{output_path}")


# ============================================================
#  FORMATTING / HIGHLIGHTING (same style as your previous)
# ============================================================
highlight_colors = {
    "Feature": "00BFCF",        # Medium Cyan
    "Area (sq. µm)": "FFEB3B",  # Medium Yellow
    "S (Wt%)": "FFA07A",        # Orange Light
    "Fe (Wt%)": "ADD8E6",       # Light Blue
    "Cu (Wt%)": "90EE90",       # Light Green
    "As (Wt%)": "D87093",       # Medium Pink
    "Ag (Wt%)": "F44336",       # Medium Red
    "Sb (Wt%)": "A9A9A9",       # Medium Grey
    "Hg (Wt%)": "9370DB",       # Medium Purple
    "Se (Wt%)": "FFA09A",
    "Te (Wt%)": "D87099",
    "Au (Wt%)": "FFF2CC",
    "Bi (Wt%)": "E2EFDA",
    "O (Wt%)": "D9E1F2",
}

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

wb = load_workbook(output_path)

# Highlight in all sheets except Area and Integrity Check
for sheet in wb.sheetnames:
    if sheet in ["Area", "Integrity Check"]:
        continue
    ws = wb[sheet]
    if ws.max_row < 2:
        continue

    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = fill

# Summary formatting (width + borders + center + bold first column)
if "Summary" in wb.sheetnames:
    ws = wb["Summary"]

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for cell in ws["A"]:
        cell.font = Font(bold=True)

wb.save(output_path)
