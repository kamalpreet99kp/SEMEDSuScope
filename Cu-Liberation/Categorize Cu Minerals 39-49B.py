import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# === Mineral classification function for Cu ===
def classify_mineral(row):
    cu = row.get("Cu (Wt%)", 0)
    s = row.get("S (Wt%)", 0)
    fe = row.get("Fe (Wt%)", 0)
    ars = row.get("As (Wt%)", 0)
    ag = row.get("Ag (Wt%)", 0)
    zn = row.get("Zn (Wt%)", 0)

    if ag >= 1.41:
        return None

    if 18.5 <= cu <= 56 and 15 <= s <= 48 and 5.0 <= ars <= 25 and fe < 27 and zn < 15:
        return "Enargite & Associations"
    elif 40 <= cu <= 85 and 10 <= s <= 45 and fe < 4 and ars < 5 and zn < 15:
        return "Covellite & Associations"
    elif 40 <= cu <= 75 and 6 <= fe <= 15 and 7 <= s <= 48 and ars < 5 and zn < 15:
        return "Bornite & Associations"
    elif 10 <= cu <= 45 and 15 <= fe <= 40 and 10 <= s <= 48 and ars < 5 and zn < 15:
        return "Chalcopyrite"
    elif cu > 2:
        return "All Others"
    else:
        return None


# === Sheet order ===
sheet_order = [
    "Area",
    "Summary",
    "Raw Data",
    "Enargite & Associations",
    "Covellite & Associations",
    "Bornite & Associations",
    "Chalcopyrite",
    "All Others",
    "Integrity Check"
]

# === File selection ===
tk.Tk().withdraw()
file_path = filedialog.askopenfilename(title="Select Full Analysis Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
if not file_path:
    print("No file selected.")
    exit()

# === Load raw data ===
df_raw = pd.read_excel(file_path, sheet_name=None)
sheet_name = list(df_raw.keys())[0]
df = df_raw[sheet_name].copy()

# === Classify data ===
df["Mineral Type"] = df.apply(classify_mineral, axis=1)

# === Prepare classified data ===
classified_sheets = {}
area_column = "Area (sq. µm)"
summary_data = []
mineral_categories = sheet_order[3:-1]

for category in mineral_categories:
    df_mineral = df[df["Mineral Type"] == category]
    classified_sheets[category] = df_mineral.drop(columns=["Mineral Type"])
    area_sum = df_mineral[area_column].sum()
    summary_data.append((category, area_sum))

# === Summary sheet ===
total_area = sum(x[1] for x in summary_data)
summary_df = pd.DataFrame(summary_data, columns=["Type of Mineral", "Total Sum of Area (sq. µm)"])
summary_df["Percentage"] = summary_df["Total Sum of Area (sq. µm)"] / total_area * 100

# === Output file path ===
output_path = os.path.join(os.path.dirname(file_path), f"Classified_{os.path.basename(file_path)}")

# === Write Excel output ===
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Area", index=False)  # Empty area sheet
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    df.drop(columns=["Mineral Type"]).to_excel(writer, sheet_name="Raw Data", index=False)
    for category in mineral_categories:
        classified_sheets[category].to_excel(writer, sheet_name=category, index=False)

    # === Integrity Check ===

    df_cu = df[(df["Cu (Wt%)"] > 2) & (df["Ag (Wt%)"] < 1.41)]
    raw_cu_count = len(df_cu)
    raw_cu_area = df_cu[area_column].sum()

    classified_total_count = sum(len(classified_sheets[cat]) for cat in mineral_categories)
    classified_total_area = sum(classified_sheets[cat][area_column].sum() for cat in mineral_categories)

    integrity_data = {
        "Metric": [
            "Rows with Cu > 2% and Ag < 1.41 in Raw Data",
            "Total rows in classified sheets",
            "Area in Raw Data (Cu > 2% and Ag < 1.41)",
            "Area in classified sheets",
            "Area Match",
            "Row Count Match"
        ],
        "Value": [
            raw_cu_count,
            classified_total_count,
            round(raw_cu_area, 4),
            round(classified_total_area, 4),
            "✅ Match" if abs(raw_cu_area - classified_total_area) < 0.01 else "❌ Mismatch",
            "✅ Match" if raw_cu_count == classified_total_count else "❌ Mismatch"
        ]
    }
    pd.DataFrame(integrity_data).to_excel(writer, sheet_name="Integrity Check", index=False)


# === Reopen and format workbook ===
wb = load_workbook(output_path)

# Color map for target columns
highlight_colors = {
    "Feature": "00BFCF",        # Medium Cyan
    "Area (sq. µm)": "FFEB3B",  # Medium Yellow
    "S (Wt%)": "FFA07A",        # Orange Light
    "Fe (Wt%)": "ADD8E6",       # Light Blue
    "Cu (Wt%)": "90EE90",       # Light Green
    "As (Wt%)": "D87093",       # Medium Pink
}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Highlight and format each sheet
for sheet in wb.sheetnames:
    ws = wb[sheet]
    if sheet == "Summary":
        # Format Summary
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        for cell in ws["A"]:
            cell.font = Font(bold=True)
        continue

    if sheet in ["Area", "Integrity Check"]:
        continue

    # Highlight target columns
    headers = {cell.value: cell.column for cell in ws[1]}
    for col_name, color in highlight_colors.items():
        if col_name in headers:
            col_idx = headers[col_name]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Save final workbook
wb.save(output_path)
print(f"\n✅ Cu Classification Complete!\n📄 Output saved to:\n{output_path}")
